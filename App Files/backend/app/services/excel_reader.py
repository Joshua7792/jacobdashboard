"""Excel workbook reader for the Cordillera dashboard.

Parses the Contractor Certifications Tracker workbook into structured data
and computes per-cell renewal status (green / yellow / red / blank), per-worker
and per-contractor compliance rollups, an urgency-sorted action list, and the
KPI bundle the dashboard needs.

Reads cached cell values via openpyxl's data_only=True so formula headers like
``=Certifications!A15`` resolve to their evaluated cert name. As a fallback
(useful if Excel hasn't cached values yet) we also resolve those references
manually by re-reading the Certifications sheet.

Color rules match the in-workbook conditional formatting:
  - GREEN  : has a date and more than 60 days remaining until 1-year anniversary
  - YELLOW : has a date and 31-60 days remaining
  - RED    : has a date and 30 days or fewer remaining (or anniversary passed)
  - BLANK  : no date entered (empty cell)
"""
from __future__ import annotations

import re
import threading
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# --- Configuration -----------------------------------------------------------

RENEWAL_RED_DAYS = 30
RENEWAL_YELLOW_DAYS = 60

CERTS_SHEET = "Certifications"
CONTRACTORS_SHEET = "Contractors"
WORKERS_SHEET = "Workers"
TRACKER_SHEET = "Tracker"

CERT_FORMULA_RE = re.compile(
    r"^\s*=\s*Certifications!\s*\$?A\$?(\d+)\s*$",
    re.IGNORECASE,
)


# --- Data structures ---------------------------------------------------------

@dataclass
class Cert:
    name: str
    category: str
    validity_years: int
    notes: Optional[str] = None


@dataclass
class CertStatus:
    cert_name: str
    cert_category: str
    completed_on: Optional[date] = None
    anniversary: Optional[date] = None
    days_until_anniversary: Optional[int] = None
    status: str = "blank"  # green | yellow | red | blank


@dataclass
class Worker:
    name: str
    contractor: str
    job_title: Optional[str] = None
    status: str = "active"
    employee_code: Optional[str] = None
    hire_date: Optional[date] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    notes: Optional[str] = None
    certs: list[CertStatus] = field(default_factory=list)
    # Computed
    compliance_pct: float = 0.0
    green_count: int = 0
    yellow_count: int = 0
    red_count: int = 0
    blank_count: int = 0


@dataclass
class Contractor:
    name: str
    primary_contact: Optional[str] = None
    specialty: Optional[str] = None
    notes: Optional[str] = None
    # Computed
    worker_count: int = 0
    compliance_pct: float = 0.0
    green_count: int = 0
    yellow_count: int = 0
    red_count: int = 0
    blank_count: int = 0
    weakest_cert: Optional[str] = None


@dataclass
class ActionItem:
    contractor: str
    worker: str
    worker_status: str
    cert_name: str
    cert_category: str
    completed_on: Optional[date]
    anniversary: Optional[date]
    days_until_anniversary: Optional[int]
    status: str  # red | yellow


@dataclass
class KPIs:
    total_contractors: int = 0
    total_workers: int = 0
    active_workers: int = 0
    total_certs: int = 0
    overall_compliance_pct: float = 0.0
    green_count: int = 0
    yellow_count: int = 0
    red_count: int = 0
    blank_count: int = 0
    today: Optional[date] = None


@dataclass
class HeatmapRow:
    worker: str
    contractor: str
    job_title: Optional[str]
    statuses: list[dict]  # [{status, completed_on, days_until_anniversary}, ...]


@dataclass
class HeatmapPayload:
    cert_names: list[str] = field(default_factory=list)
    cert_categories: list[str] = field(default_factory=list)
    rows: list[HeatmapRow] = field(default_factory=list)


@dataclass
class CertDemand:
    cert_name: str
    cert_category: str
    green: int = 0
    yellow: int = 0
    red: int = 0
    blank: int = 0
    coverage_pct: float = 0.0  # green / active_workers


@dataclass
class ParsedWorkbook:
    workbook_path: str
    last_modified: datetime
    loaded_at: datetime
    today: date
    certs: list[Cert] = field(default_factory=list)
    contractors: list[Contractor] = field(default_factory=list)
    workers: list[Worker] = field(default_factory=list)
    action_list: list[ActionItem] = field(default_factory=list)
    heatmap: HeatmapPayload = field(default_factory=HeatmapPayload)
    cert_demand: list[CertDemand] = field(default_factory=list)
    kpis: KPIs = field(default_factory=KPIs)
    issues: list[str] = field(default_factory=list)


# --- Helpers -----------------------------------------------------------------

def _normalize(s: Optional[str]) -> str:
    return (s or "").strip().lower()


def _to_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def _to_str(value) -> Optional[str]:
    if value is None or value == "":
        return None
    return str(value).strip() or None


def _anniversary(d: date, today: date) -> tuple[date, int]:
    try:
        anniv = date(d.year + 1, d.month, d.day)
    except ValueError:
        anniv = date(d.year + 1, 2, 28)
    return anniv, (anniv - today).days


def _classify(d: Optional[date], today: date) -> tuple[str, Optional[date], Optional[int]]:
    """Return (status, anniversary_date_or_None, days_remaining_or_None)."""
    if d is None:
        return "blank", None, None
    anniv, days = _anniversary(d, today)
    if days <= RENEWAL_RED_DAYS:
        return "red", anniv, days
    if days <= RENEWAL_YELLOW_DAYS:
        return "yellow", anniv, days
    return "green", anniv, days


# --- Sheet parsers -----------------------------------------------------------

def _parse_certifications(ws: Worksheet, issues: list[str]) -> list[Cert]:
    certs: list[Cert] = []
    for r in range(2, ws.max_row + 1):
        name = _to_str(ws.cell(r, 1).value)
        if name is None:
            continue
        category = _to_str(ws.cell(r, 2).value) or "Uncategorized"
        validity_raw = ws.cell(r, 3).value
        try:
            validity = int(validity_raw) if validity_raw not in (None, "") else 0
        except (TypeError, ValueError):
            validity = 0
            issues.append(
                f"Certifications!A{r}: validity '{validity_raw}' is not a number; treated as 0"
            )
        certs.append(Cert(
            name=name,
            category=category,
            validity_years=validity,
            notes=_to_str(ws.cell(r, 4).value),
        ))
    return certs


def _parse_contractors(ws: Worksheet, _issues: list[str]) -> list[Contractor]:
    contractors: list[Contractor] = []
    for r in range(2, ws.max_row + 1):
        name = _to_str(ws.cell(r, 1).value)
        if name is None:
            continue
        contractors.append(Contractor(
            name=name,
            primary_contact=_to_str(ws.cell(r, 2).value),
            specialty=_to_str(ws.cell(r, 3).value),
            notes=_to_str(ws.cell(r, 5).value),
        ))
    return contractors


def _parse_workers(ws: Worksheet, _issues: list[str]) -> list[Worker]:
    workers: list[Worker] = []
    for r in range(2, ws.max_row + 1):
        name = _to_str(ws.cell(r, 1).value)
        if name is None:
            continue
        workers.append(Worker(
            name=name,
            contractor=_to_str(ws.cell(r, 2).value) or "",
            job_title=_to_str(ws.cell(r, 3).value),
            status=_to_str(ws.cell(r, 4).value) or "active",
            employee_code=_to_str(ws.cell(r, 5).value),
            hire_date=_to_date(ws.cell(r, 6).value),
            email=_to_str(ws.cell(r, 7).value),
            phone=_to_str(ws.cell(r, 8).value),
            notes=_to_str(ws.cell(r, 9).value),
        ))
    return workers


def _parse_tracker_headers(
    ws: Worksheet,
    ws_formula: Worksheet,
    certs_lookup_by_row: dict[int, str],
    issues: list[str],
) -> list[tuple[int, str]]:
    """Return [(col, cert_name), ...] for non-empty Tracker row-2 headers.

    When the cell is a formula like =Certifications!A15, resolve it to the
    cert name in Certifications row 15. data_only=True usually does this for
    us via Excel's cached values; the formula match is the fallback.
    """
    headers: list[tuple[int, str]] = []
    for c in range(3, ws.max_column + 1):
        value = ws.cell(2, c).value
        formula_value = ws_formula.cell(2, c).value
        if value in (None, "") and isinstance(formula_value, str):
            value = formula_value
        if value is None or value == "":
            continue
        if isinstance(value, str):
            stripped = value.strip()
            m = CERT_FORMULA_RE.match(stripped)
            if m:
                ref_row = int(m.group(1))
                resolved = certs_lookup_by_row.get(ref_row)
                if resolved:
                    headers.append((c, resolved))
                else:
                    issues.append(
                        f"Tracker!{_col_letter(c)}2 references Certifications!A{ref_row} "
                        f"but that row is empty"
                    )
                continue
            if stripped.startswith("="):
                # Some other formula; openpyxl couldn't evaluate it.
                issues.append(
                    f"Tracker!{_col_letter(c)}2 contains formula '{stripped}' that "
                    f"could not be resolved (open and re-save in Excel to cache its value)"
                )
                continue
            headers.append((c, stripped))
        else:
            headers.append((c, str(value).strip()))
    return headers


def _col_letter(col: int) -> str:
    """Convert a 1-based column index to letter ('A', 'B', ..., 'AA', ...)."""
    letters = ""
    while col:
        col, rem = divmod(col - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _parse_tracker_data(
    ws: Worksheet,
    headers: list[tuple[int, str]],
    worker_lookup: dict[str, Worker],
    today: date,
    certs_by_name: dict[str, Cert],
    issues: list[str],
) -> None:
    """For each Tracker data row, attach CertStatus entries to the matching Worker."""
    for r in range(3, ws.max_row + 1):
        worker_val = ws.cell(r, 2).value
        if worker_val in (None, ""):
            continue
        worker_name = str(worker_val).strip()
        worker = worker_lookup.get(_normalize(worker_name))
        if worker is None:
            issues.append(
                f"Tracker row {r}: worker '{worker_name}' has no matching row in Workers sheet"
            )
            continue
        for col, cert_name in headers:
            cell_value = ws.cell(r, col).value
            d = _to_date(cell_value)
            if d is None and cell_value not in (None, ""):
                issues.append(
                    f"Tracker!{_col_letter(col)}{r} value '{cell_value}' could not be parsed as a date"
                )
            cert_def = certs_by_name.get(_normalize(cert_name))
            category = cert_def.category if cert_def else "Uncategorized"
            status, anniv, days = _classify(d, today)
            worker.certs.append(CertStatus(
                cert_name=cert_name,
                cert_category=category,
                completed_on=d,
                anniversary=anniv,
                days_until_anniversary=days,
                status=status,
            ))


# --- Aggregations ------------------------------------------------------------

def _aggregate_worker_counts(worker: Worker) -> None:
    counts = {"green": 0, "yellow": 0, "red": 0, "blank": 0}
    for cs in worker.certs:
        counts[cs.status] = counts.get(cs.status, 0) + 1
    worker.green_count = counts["green"]
    worker.yellow_count = counts["yellow"]
    worker.red_count = counts["red"]
    worker.blank_count = counts["blank"]
    dated = counts["green"] + counts["yellow"] + counts["red"]
    worker.compliance_pct = round(100 * counts["green"] / dated, 1) if dated > 0 else 0.0


def _aggregate_contractor_counts(contractor: Contractor, contractor_workers: list[Worker]) -> None:
    contractor.worker_count = len(contractor_workers)
    green = sum(w.green_count for w in contractor_workers)
    yellow = sum(w.yellow_count for w in contractor_workers)
    red = sum(w.red_count for w in contractor_workers)
    blank = sum(w.blank_count for w in contractor_workers)
    contractor.green_count = green
    contractor.yellow_count = yellow
    contractor.red_count = red
    contractor.blank_count = blank
    dated = green + yellow + red
    contractor.compliance_pct = round(100 * green / dated, 1) if dated > 0 else 0.0
    # Weakest cert: cert most often non-green among this contractor's workers
    cert_problems: dict[str, int] = {}
    for w in contractor_workers:
        for cs in w.certs:
            if cs.status in ("yellow", "red"):
                cert_problems[cs.cert_name] = cert_problems.get(cs.cert_name, 0) + 1
    if cert_problems:
        contractor.weakest_cert = max(cert_problems.items(), key=lambda kv: kv[1])[0]


def _build_action_list(workers: list[Worker]) -> list[ActionItem]:
    items: list[ActionItem] = []
    for w in workers:
        for cs in w.certs:
            if cs.status in ("red", "yellow") and cs.completed_on is not None:
                items.append(ActionItem(
                    contractor=w.contractor,
                    worker=w.name,
                    worker_status=w.status,
                    cert_name=cs.cert_name,
                    cert_category=cs.cert_category,
                    completed_on=cs.completed_on,
                    anniversary=cs.anniversary,
                    days_until_anniversary=cs.days_until_anniversary,
                    status=cs.status,
                ))
    # Sort by absolute distance from the anniversary boundary so the items
    # closest to needing action (a few days overdue, or about to enter the red
    # zone) bubble to the top, ahead of decades-old expirations that are
    # technically very late but functionally identical to all other reds.
    items.sort(key=lambda a: (
        abs(a.days_until_anniversary) if a.days_until_anniversary is not None else 99999
    ))
    return items


def _build_heatmap(workers: list[Worker], certs: list[Cert]) -> HeatmapPayload:
    cert_order = [c.name for c in certs]
    cert_by_name = {c.name: c for c in certs}
    rows: list[HeatmapRow] = []
    for w in sorted(workers, key=lambda x: (x.contractor, x.name)):
        cert_lookup = {cs.cert_name: cs for cs in w.certs}
        statuses: list[dict] = []
        for cn in cert_order:
            cs = cert_lookup.get(cn)
            if cs is None:
                statuses.append({
                    "status": "blank",
                    "completed_on": None,
                    "days_until_anniversary": None,
                })
            else:
                statuses.append({
                    "status": cs.status,
                    "completed_on": cs.completed_on.isoformat() if cs.completed_on else None,
                    "days_until_anniversary": cs.days_until_anniversary,
                })
        rows.append(HeatmapRow(
            worker=w.name,
            contractor=w.contractor,
            job_title=w.job_title,
            statuses=statuses,
        ))
    return HeatmapPayload(
        cert_names=cert_order,
        cert_categories=[cert_by_name[n].category for n in cert_order],
        rows=rows,
    )


def _build_cert_demand(workers: list[Worker], certs: list[Cert]) -> list[CertDemand]:
    demand: list[CertDemand] = []
    active_workers = [w for w in workers if w.status in ("active", "onboarding")]
    total = len(active_workers)
    for cert in certs:
        d = CertDemand(cert_name=cert.name, cert_category=cert.category)
        for w in active_workers:
            cs_lookup = {cs.cert_name: cs for cs in w.certs}
            cs = cs_lookup.get(cert.name)
            if cs is None or cs.status == "blank":
                d.blank += 1
            else:
                setattr(d, cs.status, getattr(d, cs.status) + 1)
        d.coverage_pct = round(100 * d.green / total, 1) if total > 0 else 0.0
        demand.append(d)
    demand.sort(key=lambda x: x.coverage_pct)  # weakest coverage first
    return demand


def _compute_kpis(
    workers: list[Worker],
    contractors: list[Contractor],
    certs: list[Cert],
    today: date,
) -> KPIs:
    green = sum(w.green_count for w in workers)
    yellow = sum(w.yellow_count for w in workers)
    red = sum(w.red_count for w in workers)
    blank = sum(w.blank_count for w in workers)
    dated = green + yellow + red
    overall = round(100 * green / dated, 1) if dated > 0 else 0.0
    return KPIs(
        total_contractors=len(contractors),
        total_workers=len(workers),
        active_workers=sum(1 for w in workers if w.status == "active"),
        total_certs=len(certs),
        overall_compliance_pct=overall,
        green_count=green,
        yellow_count=yellow,
        red_count=red,
        blank_count=blank,
        today=today,
    )


# --- Main entry --------------------------------------------------------------

def read_workbook(path: Path, today: Optional[date] = None) -> ParsedWorkbook:
    """Read the workbook and produce a ParsedWorkbook ready for the API."""
    if today is None:
        today = date.today()

    issues: list[str] = []

    try:
        wb = load_workbook(path, data_only=True, read_only=False)
        wb_formula = load_workbook(path, data_only=False, read_only=False)
    except PermissionError as exc:
        raise PermissionError(f"Workbook is locked (open in Excel): {path}") from exc

    for sheet in (CERTS_SHEET, CONTRACTORS_SHEET, WORKERS_SHEET, TRACKER_SHEET):
        if sheet not in wb.sheetnames:
            raise RuntimeError(f"Required sheet '{sheet}' not found in workbook {path.name}")

    ws_certs = wb[CERTS_SHEET]
    ws_contractors = wb[CONTRACTORS_SHEET]
    ws_workers = wb[WORKERS_SHEET]
    ws_tracker = wb[TRACKER_SHEET]
    ws_tracker_formula = wb_formula[TRACKER_SHEET]

    # Lookup for resolving =Certifications!A{n} formulas in Tracker headers.
    certs_lookup_by_row: dict[int, str] = {}
    for r in range(2, ws_certs.max_row + 1):
        v = ws_certs.cell(r, 1).value
        if v not in (None, ""):
            certs_lookup_by_row[r] = str(v).strip()

    certs = _parse_certifications(ws_certs, issues)
    contractors = _parse_contractors(ws_contractors, issues)
    workers = _parse_workers(ws_workers, issues)

    worker_lookup: dict[str, Worker] = {_normalize(w.name): w for w in workers}
    certs_by_name: dict[str, Cert] = {_normalize(c.name): c for c in certs}

    headers = _parse_tracker_headers(ws_tracker, ws_tracker_formula, certs_lookup_by_row, issues)
    _parse_tracker_data(ws_tracker, headers, worker_lookup, today, certs_by_name, issues)

    for w in workers:
        _aggregate_worker_counts(w)

    workers_by_contractor: dict[str, list[Worker]] = {}
    for w in workers:
        workers_by_contractor.setdefault(_normalize(w.contractor), []).append(w)
    for c in contractors:
        c_workers = workers_by_contractor.get(_normalize(c.name), [])
        _aggregate_contractor_counts(c, c_workers)

    action_list = _build_action_list(workers)
    heatmap = _build_heatmap(workers, certs)
    cert_demand = _build_cert_demand(workers, certs)
    kpis = _compute_kpis(workers, contractors, certs, today)

    stat = path.stat()
    return ParsedWorkbook(
        workbook_path=str(path),
        last_modified=datetime.fromtimestamp(stat.st_mtime),
        loaded_at=datetime.now(),
        today=today,
        certs=certs,
        contractors=contractors,
        workers=workers,
        action_list=action_list,
        heatmap=heatmap,
        cert_demand=cert_demand,
        kpis=kpis,
        issues=issues,
    )


# --- Cache -------------------------------------------------------------------

class WorkbookCache:
    """In-memory cache that re-reads the workbook only when its mtime changes.

    Auto-refresh on every get() if the file has been modified since the cached
    snapshot was loaded. The /api/excel/refresh endpoint forces a reload via
    `force=True` regardless of mtime.
    """

    def __init__(self, path_resolver: Callable[[], Path]):
        self._path_resolver = path_resolver
        self._data: Optional[ParsedWorkbook] = None
        self._mtime: float = 0.0
        self._lock = threading.Lock()

    def get(self, force: bool = False) -> ParsedWorkbook:
        with self._lock:
            path = self._path_resolver()
            if not path.exists():
                raise FileNotFoundError(f"Workbook not found: {path}")
            current_mtime = path.stat().st_mtime
            if force or self._data is None or current_mtime > self._mtime:
                self._data = read_workbook(path)
                self._mtime = current_mtime
            return self._data

    def health(self) -> dict:
        path = self._path_resolver()
        info: dict = {
            "workbook_path": str(path),
            "exists": path.exists(),
            "loaded": self._data is not None,
        }
        if path.exists():
            info["last_modified"] = datetime.fromtimestamp(path.stat().st_mtime).isoformat()
        if self._data is not None:
            info["last_loaded"] = self._data.loaded_at.isoformat()
            info["issues_count"] = len(self._data.issues)
            info["today"] = self._data.today.isoformat()
        return info
