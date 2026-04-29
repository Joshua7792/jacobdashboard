"""Workbook structure sync for manual certification edits.

Tracker is the manual fallback source for new certification columns. When a
user manually adds a new Tracker header, this module adds the corresponding
Certifications row and formula-links that Tracker header back to the new row.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet


CERT_FORMULA_RE = re.compile(
    r"^\s*=\s*Certifications!\s*\$?A\$?(\d+)\s*$",
    re.IGNORECASE,
)
DASHBOARD_DATA_START_ROW = 9


def _normalize(s: object) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _first_empty_row(ws: Worksheet, key_col: int, start: int = 2) -> int:
    for row in range(start, max(ws.max_row + 50, 2000)):
        if ws.cell(row, key_col).value in (None, ""):
            return row
    return ws.max_row + 1


def _get_certifications(ws_certs: Worksheet) -> list[tuple[int, str, str]]:
    certs: list[tuple[int, str, str]] = []
    for row in range(2, ws_certs.max_row + 1):
        name = ws_certs.cell(row, 1).value
        if name in (None, ""):
            continue
        category = ws_certs.cell(row, 2).value or "Additional Training"
        certs.append((row, str(name), str(category)))
    return certs


def _tracker_column_for_cert_row(ws_tracker: Worksheet, certs_row: int) -> int | None:
    for col in range(3, ws_tracker.max_column + 1):
        value = ws_tracker.cell(2, col).value
        if isinstance(value, str):
            match = CERT_FORMULA_RE.match(value)
            if match and int(match.group(1)) == certs_row:
                return col
    return None


def _tracker_literal_column(ws_tracker: Worksheet, cert_name: str) -> int | None:
    target = _normalize(cert_name)
    for col in range(3, ws_tracker.max_column + 1):
        value = ws_tracker.cell(2, col).value
        if not isinstance(value, str) or value.startswith("="):
            continue
        if _normalize(value) == target:
            return col
    return None


def _apply_renewal_color_rules(ws_tracker: Worksheet) -> None:
    last_row = max(ws_tracker.max_row, 200)
    last_col = max(ws_tracker.max_column, 3)
    matrix_range = f"C3:{get_column_letter(last_col)}{last_row}"

    ws_tracker.conditional_formatting._cf_rules.clear()

    red_fill = PatternFill(bgColor="F8CBAD", fill_type="solid")
    yellow_fill = PatternFill(bgColor="FFE699", fill_type="solid")
    green_fill = PatternFill(bgColor="C6EFCE", fill_type="solid")

    ws_tracker.conditional_formatting.add(
        matrix_range,
        FormulaRule(
            formula=['AND(C3<>"",ISNUMBER(C3),(EDATE(C3,12)-TODAY())<=30)'],
            fill=red_fill,
            stopIfTrue=True,
        ),
    )
    ws_tracker.conditional_formatting.add(
        matrix_range,
        FormulaRule(
            formula=[
                'AND(C3<>"",ISNUMBER(C3),'
                '(EDATE(C3,12)-TODAY())>30,'
                '(EDATE(C3,12)-TODAY())<=60)'
            ],
            fill=yellow_fill,
            stopIfTrue=True,
        ),
    )
    ws_tracker.conditional_formatting.add(
        matrix_range,
        FormulaRule(
            formula=['AND(C3<>"",ISNUMBER(C3),(EDATE(C3,12)-TODAY())>60)'],
            fill=green_fill,
            stopIfTrue=True,
        ),
    )


def _sync_tracker(ws_certs: Worksheet, ws_tracker: Worksheet) -> dict[str, list[str]]:
    actions = {
        "new_certs_from_tracker": [],
        "converted_tracker_headers": [],
    }

    certs = _get_certifications(ws_certs)
    known = {_normalize(name) for _row, name, _category in certs}

    for col in range(3, ws_tracker.max_column + 1):
        value = ws_tracker.cell(2, col).value
        if not isinstance(value, str) or not value or value.startswith("="):
            continue
        if _normalize(value) in known:
            continue
        row = _first_empty_row(ws_certs, key_col=1, start=2)
        ws_certs.cell(row, 1, value=value)
        ws_certs.cell(row, 2, value="Additional Training")
        ws_certs.cell(row, 3, value=0)
        actions["new_certs_from_tracker"].append(value)

    certs = _get_certifications(ws_certs)
    for certs_row, name, _category in certs:
        formula_col = _tracker_column_for_cert_row(ws_tracker, certs_row)
        if formula_col is not None:
            continue
        literal_col = _tracker_literal_column(ws_tracker, name)
        if literal_col is not None:
            ws_tracker.cell(2, literal_col, value=f"=Certifications!A{certs_row}")
            actions["converted_tracker_headers"].append(name)

    last_row = max(ws_tracker.max_row, 200)
    last_col = max(ws_tracker.max_column, 26)
    ws_tracker.auto_filter.ref = f"A2:{get_column_letter(last_col)}{last_row}"
    _apply_renewal_color_rules(ws_tracker)

    return actions


def _sync_dashboard(ws_certs: Worksheet, ws_dashboard: Worksheet) -> dict[str, list[str]]:
    actions = {"added_dashboard_rows": [], "converted_dashboard_rows": []}
    certs = _get_certifications(ws_certs)

    by_cert_row: dict[int, int] = {}
    literal_rows: dict[str, int] = {}
    used_rows: set[int] = set()
    end_row = max(DASHBOARD_DATA_START_ROW, ws_dashboard.max_row) + 5

    for row in range(DASHBOARD_DATA_START_ROW, end_row):
        value = ws_dashboard.cell(row, 1).value
        if value in (None, ""):
            continue
        used_rows.add(row)
        if isinstance(value, str):
            match = CERT_FORMULA_RE.match(value)
            if match:
                by_cert_row[int(match.group(1))] = row
            else:
                literal_rows[_normalize(value)] = row

    def next_free_row() -> int:
        row = DASHBOARD_DATA_START_ROW
        while row in used_rows:
            row += 1
        used_rows.add(row)
        return row

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for certs_row, name, _category in certs:
        is_new = False
        if certs_row in by_cert_row:
            dashboard_row = by_cert_row[certs_row]
        elif _normalize(name) in literal_rows:
            dashboard_row = literal_rows[_normalize(name)]
            actions["converted_dashboard_rows"].append(name)
        else:
            dashboard_row = next_free_row()
            is_new = True
            actions["added_dashboard_rows"].append(name)

        row = dashboard_row
        ws_dashboard.cell(row, 1, value=f"=Certifications!A{certs_row}")
        ws_dashboard.cell(row, 2, value=f"=Certifications!B{certs_row}")
        ws_dashboard.cell(
            row,
            3,
            value=(
                f"=IFERROR(COUNTA(INDEX(Tracker!$C$3:$AZ$200,0,"
                f"MATCH($A{row},Tracker!$C$2:$AZ$2,0))),0)"
            ),
        )
        ws_dashboard.cell(row, 4, value=f"=COUNTA(Workers!A$2:A$200)-C{row}")
        pct = ws_dashboard.cell(row, 5, value=f"=IFERROR(C{row}/(C{row}+D{row}),0)")
        pct.number_format = "0.0%"
        bar = ws_dashboard.cell(row, 6, value=f"=REPT(CHAR(9632),ROUND(E{row}*20,0))")
        if is_new:
            bar.font = Font(color="2E7D32", size=11)
            for col in range(1, 7):
                ws_dashboard.cell(row, col).border = border

    return actions


def _sync_certifications_table(ws_certs: Worksheet) -> list[str]:
    if "CertificationsTable" not in ws_certs.tables:
        return []
    table = ws_certs.tables["CertificationsTable"]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    target_row = max(ws_certs.max_row, max_row, 60)
    if target_row <= max_row:
        return []
    table.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{target_row}"
    )
    return [f"CertificationsTable now covers row {target_row}"]


def sync_workbook_file(path: Path) -> dict[str, list[str]]:
    """Sync the workbook in-place and return a small change summary."""
    workbook = load_workbook(path)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    required = {"Certifications", "Tracker", "Dashboard"}
    missing = required - set(workbook.sheetnames)
    if missing:
        raise RuntimeError(f"Workbook is missing sheet(s): {', '.join(sorted(missing))}")

    ws_certs = workbook["Certifications"]
    ws_tracker = workbook["Tracker"]
    ws_dashboard = workbook["Dashboard"]

    actions: dict[str, list[str]] = {}
    for prefix, result in (
        ("tracker", _sync_tracker(ws_certs, ws_tracker)),
        ("dashboard", _sync_dashboard(ws_certs, ws_dashboard)),
    ):
        for key, values in result.items():
            if values:
                actions[f"{prefix}.{key}"] = values

    table_actions = _sync_certifications_table(ws_certs)
    if table_actions:
        actions["certifications.table_extended"] = table_actions

    if actions:
        workbook.save(path)
    return actions
