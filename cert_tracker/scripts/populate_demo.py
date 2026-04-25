"""Populate the Demo workbook with dummy contractors, workers, and cert dates.

Designed to exercise every Tracker conditional-formatting state:
  - green  = current (date within validity window)
  - yellow = expiring within 60 days
  - orange = expired (past validity)
  - red    = missing (blank cell under an existing worker)

Safe by default: refuses to overwrite if the demo workbook already contains
contractor or worker rows. Pass --force to clear and repopulate.

Usage:
    python scripts/populate_demo.py           # refuses if data exists
    python scripts/populate_demo.py --force   # wipe and repopulate
"""
from __future__ import annotations

import random
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "Contractor Certifications Tracker Demo.xlsx"

# Reference date used when designing these slots: 2026-04-23.
CURRENT_1YR = date(2025, 11, 15)
CURRENT_3YR = date(2024, 8, 10)
EXPIRING_1YR = date(2025, 5, 20)
EXPIRING_3YR = date(2023, 5, 20)
EXPIRED_1YR = date(2024, 6, 15)
EXPIRED_3YR = date(2022, 11, 10)

C = CURRENT_1YR
C3 = CURRENT_3YR
Y = EXPIRING_1YR
Y3 = EXPIRING_3YR
E = EXPIRED_1YR
E3 = EXPIRED_3YR

# Validity years per cert (matches Certifications sheet).
VALIDITY_YEARS: dict[str, int] = {
    "Protección Contra Caídas": 1,
    "Flama Expuesta": 1,
    "Espacios Confinados": 1,
    "Manejo de Equipo Motorizado": 3,
    "Excavación o Zanja": 1,
    "Manejo de Tijeras (Scissor Lift)": 3,
    "Seguridad Eléctrica": 1,
    "Lockout": 1,
    "Manejo de Grúas": 3,
    "Escaleras": 3,
    "Andamios": 3,
    "Trabajos con Plomo o Asbesto": 1,
    "Comunicación de Riesgos": 1,
    "Inducción Jacobs/Lilly": 1,
    "OSHA 40Hr HAZWOPER": 0,
    "OSHA 8Hr Refresher": 1,
    "Drilling Safety": 3,
    "Utility Locating": 1,
    "OSHA 10": 0,
    "OSHA 30": 0,
    "Rebar Safety": 0,
    "Formwork & Shoring": 0,
    "Silica Exposure": 0,
    "Concrete & Masonry": 0,
}
ALL_CERTS = list(VALIDITY_YEARS.keys())

# Maps free-form cert names (as typed in the second dummy-data batch) to the
# canonical names used on the Certifications sheet. Names mapped to None have
# no equivalent in the catalog and are dropped; random fill covers that worker.
CERT_ALIAS_FOR_INPUT: dict[str, str | None] = {
    "OSHA 10": "OSHA 10",
    "OSHA 30": "OSHA 30",
    "Silica Exposure": "Silica Exposure",
    "Fall Protection": "Protección Contra Caídas",
    "Confined Space": "Espacios Confinados",
    "Lockout/Tagout": "Lockout",
    "Excavation Safety": "Excavación o Zanja",
    "Concrete Safety": "Concrete & Masonry",
    "First Aid": None,
    "Steel Erection": None,
    "Traffic Control": None,
}

# --- Dataset 1: hand-curated contractors/workers/certs ---
CONTRACTORS_CURATED = [
    ("Stonewall Drilling Services, LLC", "Juan Pérez Vázquez", "Drilling / Geotechnical"),
    ("Atlas Concrete Works, Inc.", "Carla M. Santos Vega", "Concrete & Foundation"),
    ("PowerLine Electric Corp.", "Miguel Torres Ramos", "Electrical & Instrumentation"),
    ("SafePath Excavation, Co.", "Ana L. Reyes Díaz", "Excavation & Earthworks"),
    ("SkyHigh Scaffolding Solutions", "Roberto Díaz Cruz", "Scaffolding / Access Systems"),
]

# Each curated worker: (name, title, status, hire, contractor, cert_map)
# employee_code / email / phone / notes default to empty strings.
WORKERS_CURATED = [
    ("Pedro A. Maldonado", "Lead Driller", "active", date(2021, 3, 15),
     "Stonewall Drilling Services, LLC",
     {"Protección Contra Caídas": C, "Flama Expuesta": C, "Espacios Confinados": C,
      "Manejo de Equipo Motorizado": C3, "Excavación o Zanja": C,
      "Manejo de Tijeras (Scissor Lift)": C3, "Seguridad Eléctrica": C, "Lockout": C,
      "Manejo de Grúas": C3, "Escaleras": C3, "Andamios": C3,
      "Trabajos con Plomo o Asbesto": C, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 40Hr HAZWOPER": date(2015, 6, 1),
      "OSHA 8Hr Refresher": C, "Drilling Safety": C3, "Utility Locating": C}),
    ("Samuel Ortiz Burgos", "Driller Helper", "active", date(2023, 7, 10),
     "Stonewall Drilling Services, LLC",
     {"Protección Contra Caídas": Y, "Flama Expuesta": Y, "Espacios Confinados": Y,
      "Manejo de Equipo Motorizado": Y3, "Excavación o Zanja": Y,
      "Manejo de Tijeras (Scissor Lift)": Y3, "Seguridad Eléctrica": Y, "Lockout": Y,
      "Escaleras": Y3, "Andamios": Y3,
      "Trabajos con Plomo o Asbesto": Y, "Comunicación de Riesgos": Y,
      "Inducción Jacobs/Lilly": Y, "OSHA 40Hr HAZWOPER": date(2020, 4, 10),
      "OSHA 8Hr Refresher": Y, "Drilling Safety": Y3, "Utility Locating": Y}),
    ("Luis R. Negrón Ruiz", "Driller", "active", date(2022, 1, 25),
     "Stonewall Drilling Services, LLC",
     {"Protección Contra Caídas": C, "Flama Expuesta": C, "Espacios Confinados": C,
      "Manejo de Equipo Motorizado": C3, "Excavación o Zanja": C,
      "Seguridad Eléctrica": C, "Lockout": C, "Manejo de Grúas": C3,
      "Andamios": C3, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 40Hr HAZWOPER": date(2018, 9, 1),
      "OSHA 8Hr Refresher": E, "Drilling Safety": C3}),
    ("Marta I. Colón", "Apprentice", "onboarding", date(2026, 2, 1),
     "Stonewall Drilling Services, LLC",
     {"Protección Contra Caídas": C, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 10": date(2025, 10, 1)}),
    ("Raúl García Molina", "Foreman", "active", date(2019, 5, 20),
     "Atlas Concrete Works, Inc.",
     {"Protección Contra Caídas": C, "Flama Expuesta": C, "Espacios Confinados": C,
      "Manejo de Equipo Motorizado": C3, "Excavación o Zanja": C,
      "Manejo de Tijeras (Scissor Lift)": C3, "Seguridad Eléctrica": C, "Lockout": C,
      "Manejo de Grúas": C3, "Escaleras": C3, "Andamios": C3,
      "Trabajos con Plomo o Asbesto": C, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 30": date(2019, 7, 22),
      "Rebar Safety": date(2022, 3, 14), "Formwork & Shoring": date(2023, 9, 10),
      "Silica Exposure": date(2024, 2, 5), "Concrete & Masonry": date(2023, 11, 18)}),
    ("Ismael Villa", "Concrete Finisher", "active", date(2020, 8, 15),
     "Atlas Concrete Works, Inc.",
     {"Protección Contra Caídas": C, "Flama Expuesta": E, "Espacios Confinados": C,
      "Manejo de Equipo Motorizado": C3,
      "Manejo de Tijeras (Scissor Lift)": E3, "Seguridad Eléctrica": C, "Lockout": C,
      "Escaleras": C3, "Andamios": C3, "Trabajos con Plomo o Asbesto": C,
      "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 10": date(2022, 1, 20),
      "Rebar Safety": date(2021, 8, 10), "Silica Exposure": date(2024, 4, 4),
      "Concrete & Masonry": date(2024, 6, 12)}),
    ("Diego Acosta", "Laborer", "active", date(2024, 11, 1),
     "Atlas Concrete Works, Inc.",
     {"Protección Contra Caídas": C, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": Y, "OSHA 10": date(2024, 11, 5)}),
    ("Javier Méndez Torres", "Master Electrician", "active", date(2015, 4, 4),
     "PowerLine Electric Corp.",
     {"Protección Contra Caídas": C, "Flama Expuesta": C, "Espacios Confinados": C,
      "Manejo de Equipo Motorizado": Y3,
      "Manejo de Tijeras (Scissor Lift)": C3, "Seguridad Eléctrica": C, "Lockout": C,
      "Escaleras": C3, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 30": date(2017, 3, 22),
      "OSHA 8Hr Refresher": C}),
    ("Óscar R. Figueroa", "Electrician", "active", date(2021, 9, 8),
     "PowerLine Electric Corp.",
     {"Protección Contra Caídas": E, "Flama Expuesta": C, "Espacios Confinados": C,
      "Seguridad Eléctrica": C, "Lockout": E,
      "Escaleras": C3, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 10": date(2020, 5, 1),
      "OSHA 8Hr Refresher": Y}),
    ("Edwin Ruiz Pagán", "Apprentice Electrician", "onboarding", date(2025, 10, 10),
     "PowerLine Electric Corp.",
     {"Protección Contra Caídas": C, "Seguridad Eléctrica": Y,
      "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 10": date(2025, 11, 30)}),
    ("Antonio Benítez", "Equipment Operator", "active", date(2018, 6, 20),
     "SafePath Excavation, Co.",
     {"Protección Contra Caídas": C, "Flama Expuesta": C, "Espacios Confinados": C,
      "Manejo de Equipo Motorizado": C3, "Excavación o Zanja": C,
      "Manejo de Tijeras (Scissor Lift)": C3, "Seguridad Eléctrica": C, "Lockout": C,
      "Manejo de Grúas": C3, "Escaleras": C3,
      "Trabajos con Plomo o Asbesto": C, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 30": date(2018, 2, 12),
      "Utility Locating": C}),
    ("Fernando Soto", "Operator", "active", date(2022, 4, 15),
     "SafePath Excavation, Co.",
     {"Protección Contra Caídas": Y, "Flama Expuesta": Y, "Espacios Confinados": Y,
      "Manejo de Equipo Motorizado": Y3, "Excavación o Zanja": Y,
      "Manejo de Tijeras (Scissor Lift)": Y3, "Seguridad Eléctrica": Y, "Lockout": Y,
      "Manejo de Grúas": Y3, "Comunicación de Riesgos": Y,
      "Inducción Jacobs/Lilly": Y, "OSHA 30": date(2021, 9, 1),
      "Utility Locating": Y}),
    ("Rafael Martínez", "Spotter", "active", date(2023, 2, 28),
     "SafePath Excavation, Co.",
     {"Protección Contra Caídas": E, "Flama Expuesta": C, "Espacios Confinados": C,
      "Excavación o Zanja": E, "Seguridad Eléctrica": C,
      "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 10": date(2023, 6, 14)}),
    ("Gabriel Estévez", "Scaffold Erector", "active", date(2017, 10, 11),
     "SkyHigh Scaffolding Solutions",
     {"Protección Contra Caídas": C, "Flama Expuesta": C, "Espacios Confinados": C,
      "Manejo de Equipo Motorizado": C3,
      "Manejo de Tijeras (Scissor Lift)": C3, "Seguridad Eléctrica": C, "Lockout": C,
      "Escaleras": C3, "Andamios": C3, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C, "OSHA 30": date(2019, 11, 11)}),
    ("Mateo Villanueva", "Scaffold Helper", "active", date(2023, 1, 17),
     "SkyHigh Scaffolding Solutions",
     {"Protección Contra Caídas": Y, "Flama Expuesta": C, "Espacios Confinados": C,
      "Manejo de Tijeras (Scissor Lift)": C3, "Seguridad Eléctrica": C, "Lockout": C,
      "Escaleras": C3, "Andamios": C3, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": Y, "OSHA 10": date(2024, 3, 20)}),
    ("Leonardo Benítez", "Apprentice", "onboarding", date(2026, 3, 5),
     "SkyHigh Scaffolding Solutions",
     {"Protección Contra Caídas": C, "Comunicación de Riesgos": C,
      "Inducción Jacobs/Lilly": C}),
]


# --- Dataset 2: user-provided dummy data (notes column kept empty) ---
CONTRACTORS_USER = [
    ("Atlantic Safety Builders, LLC", "Maria Santos", "General Construction",
     "Demo contractor"),
    ("North Shore Mechanical, Inc.", "Carlos Rivera", "Mechanical / Utilities",
     "Demo contractor"),
    ("Caribbean Concrete Works", "Andrea Cruz", "Concrete and Masonry",
     "Demo contractor"),
    ("IronCore Contractors", "Daniel Perez", "Structural Steel",
     "Demo contractor"),
    ("Coastal Infrastructure LLC", "Miguel Lopez", "Civil / Infrastructure",
     "Demo contractor"),
    ("Pioneer Site Solutions", "Anthony Ramos", "Site Development",
     "Demo contractor"),
]

# Each entry: (name, contractor, title, status, code, hire, email, phone, notes,
#              seed_certs list of (raw_cert_name, date))
WORKERS_USER_RAW = [
    ("John Martinez", "Atlantic Safety Builders, LLC", "Foreman", "active",
     "ASB-001", date(2022, 3, 15), "john.martinez@asb.com", "787-555-0101", "",
     [("OSHA 10", date(2025, 1, 15)), ("Fall Protection", date(2025, 2, 20))]),
    ("Luis Gonzalez", "Atlantic Safety Builders, LLC", "Laborer", "active",
     "ASB-002", date(2023, 7, 22), "luis.gonzalez@asb.com", "787-555-0102", "",
     [("OSHA 30", date(2025, 3, 10))]),
    ("Maria Torres", "Atlantic Safety Builders, LLC", "Safety Officer", "inactive",
     "ASB-003", date(2021, 11, 5), "maria.torres@asb.com", "787-555-0103", "",
     [("First Aid", date(2024, 11, 5))]),
    ("Carlos Rivera", "North Shore Mechanical, Inc.", "Supervisor", "active",
     "NSM-001", date(2020, 6, 10), "carlos.rivera@nsm.com", "787-555-0201", "",
     [("OSHA 30", date(2025, 1, 12))]),
    ("Javier Morales", "North Shore Mechanical, Inc.", "Technician", "active",
     "NSM-002", date(2022, 9, 18), "javier.morales@nsm.com", "787-555-0202", "",
     [("Confined Space", date(2025, 2, 18))]),
    ("Elena Ruiz", "North Shore Mechanical, Inc.", "Engineer", "active",
     "NSM-003", date(2024, 1, 12), "elena.ruiz@nsm.com", "787-555-0203", "",
     [("Lockout/Tagout", date(2025, 3, 22))]),
    ("Andrea Cruz", "Caribbean Concrete Works", "Project Manager", "active",
     "CCW-001", date(2021, 4, 25), "andrea.cruz@ccw.com", "787-555-0301", "",
     [("Concrete Safety", date(2025, 1, 30))]),
    ("Pedro Santiago", "Caribbean Concrete Works", "Mason", "active",
     "CCW-002", date(2023, 2, 14), "pedro.santiago@ccw.com", "787-555-0302", "",
     [("OSHA 10", date(2025, 2, 14))]),
    ("Ricardo Vega", "Caribbean Concrete Works", "Laborer", "inactive",
     "CCW-003", date(2020, 8, 30), "ricardo.vega@ccw.com", "787-555-0303", "",
     [("Silica Exposure", date(2024, 12, 1))]),
    ("Daniel Perez", "IronCore Contractors", "Welder", "active",
     "ICC-001", date(2022, 5, 19), "daniel.perez@ironcore.com", "787-555-0401", "",
     [("Steel Erection", date(2025, 1, 25))]),
    ("Sofia Martinez", "IronCore Contractors", "Inspector", "active",
     "ICC-002", date(2023, 10, 3), "sofia.martinez@ironcore.com", "787-555-0402", "",
     [("OSHA 10", date(2025, 3, 5))]),
    ("Miguel Lopez", "Coastal Infrastructure LLC", "Civil Engineer", "active",
     "CIL-001", date(2021, 1, 11), "miguel.lopez@coastal.com", "787-555-0501", "",
     [("Excavation Safety", date(2025, 2, 8))]),
    ("Valeria Ortiz", "Coastal Infrastructure LLC", "Surveyor", "active",
     "CIL-002", date(2024, 2, 20), "valeria.ortiz@coastal.com", "787-555-0502", "",
     [("Traffic Control", date(2025, 3, 12))]),
    ("Anthony Ramos", "Pioneer Site Solutions", "Site Manager", "active",
     "PSS-001", date(2022, 12, 8), "anthony.ramos@pss.com", "787-555-0601", "",
     [("OSHA 30", date(2025, 1, 18))]),
    ("Camila Reyes", "Pioneer Site Solutions", "Coordinator", "inactive",
     "PSS-002", date(2021, 6, 27), "camila.reyes@pss.com", "787-555-0602", "",
     [("First Aid", date(2024, 10, 22))]),
]


def _random_date_for_cert(cert_name: str, rng: random.Random) -> date | None:
    """Pick a random date/state for a cert, biased toward current."""
    years = VALIDITY_YEARS[cert_name]
    state = rng.choices(
        ["current", "current", "current", "expiring", "expired", "missing"],
        k=1,
    )[0]
    if state == "missing":
        return None
    if years == 0:
        return date(rng.randint(2018, 2025), rng.randint(1, 12), rng.randint(1, 28))
    slot = {
        (1, "current"): CURRENT_1YR,
        (3, "current"): CURRENT_3YR,
        (1, "expiring"): EXPIRING_1YR,
        (3, "expiring"): EXPIRING_3YR,
        (1, "expired"): EXPIRED_1YR,
        (3, "expired"): EXPIRED_3YR,
    }
    return slot.get((years, state))


def _build_cert_dict(
    seed_certs: list[tuple[str, date]], status: str, rng: random.Random
) -> dict[str, date]:
    """Start with seed certs (alias-mapped), then fill with random extras."""
    certs: dict[str, date] = {}
    for raw_name, raw_date in seed_certs:
        mapped = CERT_ALIAS_FOR_INPUT.get(raw_name, raw_name)
        if mapped and mapped in VALIDITY_YEARS:
            certs[mapped] = raw_date

    # Target cert count by status — apprentices/inactive workers have fewer.
    if status == "onboarding":
        target = rng.randint(3, 8)
    elif status == "inactive":
        target = rng.randint(3, 10)
    else:
        target = rng.randint(6, 18)

    pool = [c for c in ALL_CERTS if c not in certs]
    rng.shuffle(pool)
    for cert_name in pool:
        if len(certs) >= target:
            break
        dt = _random_date_for_cert(cert_name, rng)
        if dt is not None:
            certs[cert_name] = dt
    return certs


def _iter_workers(rng: random.Random):
    """Yield (name, title, status, hire, contractor, certs, code, email, phone, notes)."""
    for name, title, status, hire, contractor, certs in WORKERS_CURATED:
        yield (name, title, status, hire, contractor, certs, "", "", "", "")
    for (name, contractor, title, status, code, hire, email, phone, notes,
         seed_certs) in WORKERS_USER_RAW:
        certs = _build_cert_dict(seed_certs, status, rng)
        yield (name, title, status, hire, contractor, certs, code, email, phone, notes)


def _iter_contractors():
    for name, contact, specialty in CONTRACTORS_CURATED:
        yield (name, contact, specialty, "")
    for name, contact, specialty, notes in CONTRACTORS_USER:
        yield (name, contact, specialty, notes)


def has_data(ws, col: int = 1) -> bool:
    return ws.cell(row=2, column=col).value not in (None, "")


def clear_rows(ws, start_row: int, num_cols: int, skip_cols: set[int] | None = None) -> None:
    skip_cols = skip_cols or set()
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, num_cols + 1):
            if c in skip_cols:
                continue
            ws.cell(row=r, column=c).value = None


def main(argv: list[str] | None = None) -> int:
    args = list(argv if argv is not None else sys.argv[1:])
    force = "--force" in args

    if not WORKBOOK.exists():
        print(f"Demo workbook not found: {WORKBOOK}")
        return 1

    try:
        wb = load_workbook(WORKBOOK)
    except PermissionError:
        print("[ERROR] Demo workbook is open in Excel. Close it and try again.")
        return 2

    ws_contractors = wb["Contractors"]
    ws_workers = wb["Workers"]
    ws_tracker = wb["Tracker"]

    if (has_data(ws_contractors) or has_data(ws_workers)) and not force:
        print("=" * 70)
        print("SAFETY STOP: Demo workbook already has contractor or worker data.")
        print(f"  {WORKBOOK}")
        print()
        print("Pass --force to wipe Contractors/Workers/Tracker rows and repopulate")
        print("with the dummy dataset.")
        print("=" * 70)
        return 3

    if force:
        clear_rows(ws_contractors, start_row=2, num_cols=5, skip_cols={4})
        clear_rows(ws_workers, start_row=2, num_cols=9)
        clear_rows(ws_tracker, start_row=3, num_cols=26)

    # Map Tracker cert name -> column index by reading row 2 headers
    cert_col_map: dict[str, int] = {}
    for c in range(3, ws_tracker.max_column + 1):
        v = ws_tracker.cell(row=2, column=c).value
        if v:
            cert_col_map[str(v).strip()] = c

    # Populate Contractors
    contractors = list(_iter_contractors())
    for i, (name, contact, specialty, notes) in enumerate(contractors, start=2):
        ws_contractors.cell(row=i, column=1, value=name)
        ws_contractors.cell(row=i, column=2, value=contact)
        ws_contractors.cell(row=i, column=3, value=specialty)
        count_cell = ws_contractors.cell(row=i, column=4)
        if not count_cell.value:
            count_cell.value = f'=IF(A{i}="","",COUNTIF(Workers!B:B,A{i}))'
        ws_contractors.cell(row=i, column=5, value=notes or None)

    # Populate Workers + Tracker
    rng = random.Random(42)  # reproducible random fill
    unknown_certs: set[str] = set()
    total_dates = 0
    workers = list(_iter_workers(rng))
    for i, (name, title, status, hire, contractor, certs, code, email, phone, notes) \
            in enumerate(workers, start=2):
        ws_workers.cell(row=i, column=1, value=name)
        ws_workers.cell(row=i, column=2, value=contractor)
        ws_workers.cell(row=i, column=3, value=title)
        ws_workers.cell(row=i, column=4, value=status)
        if code:
            ws_workers.cell(row=i, column=5, value=code)
        hire_cell = ws_workers.cell(row=i, column=6, value=hire)
        hire_cell.number_format = "mm/dd/yyyy"
        if email:
            ws_workers.cell(row=i, column=7, value=email)
        if phone:
            ws_workers.cell(row=i, column=8, value=phone)
        if notes:
            ws_workers.cell(row=i, column=9, value=notes)

        t_row = i + 1  # Tracker row 3 aligns with Workers row 2
        ws_tracker.cell(row=t_row, column=1, value=contractor)
        ws_tracker.cell(row=t_row, column=2, value=name)
        for cert_name, dt in certs.items():
            col = cert_col_map.get(cert_name)
            if col is None:
                unknown_certs.add(cert_name)
                continue
            cell = ws_tracker.cell(row=t_row, column=col, value=dt)
            cell.number_format = "mm/dd/yyyy"
            total_dates += 1

    wb.save(WORKBOOK)

    print(f"Saved: {WORKBOOK}")
    print(f"  Contractors:  {len(contractors)}")
    print(f"  Workers:      {len(workers)}")
    print(f"  Cert dates:   {total_dates}")
    if unknown_certs:
        print("  [warn] Unknown cert columns (skipped):")
        for name in sorted(unknown_certs):
            print(f"         - {name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
