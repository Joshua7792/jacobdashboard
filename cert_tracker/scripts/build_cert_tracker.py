"""Build the Contractor Certification Tracker workbook.

Generates `Contractor Certifications Tracker.xlsx` in the project root.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "Contractor Certifications Tracker.xlsx"

HSE_REQUIRED = [
    ("Protección Contra Caídas", 1),
    ("Flama Expuesta", 1),
    ("Espacios Confinados", 1),
    ("Manejo de Equipo Motorizado", 3),
    ("Excavación o Zanja", 1),
    ("Manejo de Tijeras (Scissor Lift)", 3),
    ("Seguridad Eléctrica", 1),
    ("Lockout", 1),
    ("Manejo de Grúas", 3),
    ("Escaleras", 3),
    ("Andamios", 3),
    ("Trabajos con Plomo o Asbesto", 1),
    ("Comunicación de Riesgos", 1),
]

ADDITIONAL_TRAINING = [
    ("Inducción Jacobs/Lilly", 1),
    ("OSHA 40Hr HAZWOPER", 0),  # 0 = initial/no expiration, refresher handles it
    ("OSHA 8Hr Refresher", 1),
    ("Drilling Safety", 3),
    ("Utility Locating", 1),
    ("OSHA 10", 0),
    ("OSHA 30", 0),
    ("Rebar Safety", 0),
    ("Formwork & Shoring", 0),
    ("Silica Exposure", 0),
    ("Concrete & Masonry", 0),
]

ALL_CERTS = [(n, "HSE Required", v) for n, v in HSE_REQUIRED] + [
    (n, "Additional Training", v) for n, v in ADDITIONAL_TRAINING
]

CONTRACTORS = [
    ("GeoEnviroTech, Inc.", "Juan D. Negrón, PG", "Drilling / Geotechnical", ""),
]

WORKERS = [
    ("William Rodriguez Rivera", "GeoEnviroTech, Inc.", "Driller", "active"),
    ("Hardy Rodríguez Vázquez", "GeoEnviroTech, Inc.", "Driller", "active"),
    ("Juan D. Negrón Hernández", "GeoEnviroTech, Inc.", "Project Geologist", "active"),
    ("Rafael Díaz Colón", "GeoEnviroTech, Inc.", "Driller", "active"),
    ("Luis A. Padilla Díaz", "GeoEnviroTech, Inc.", "Driller Helper", "active"),
    ("Jesús Quiñonez Ayala", "GeoEnviroTech, Inc.", "Driller Helper", "active"),
]

D = lambda m, d, y: date(y, m, d)  # noqa: E731
# Cert data pulled from the uploaded PDF (Anejo 3, completed 04/08/2026)
CERT_DATA = {
    "William Rodriguez Rivera": {
        "Protección Contra Caídas": D(4, 17, 2025),
        "Seguridad Eléctrica": D(4, 17, 2025),
        "Lockout": D(4, 17, 2025),
        "Comunicación de Riesgos": D(4, 17, 2025),
        "OSHA 40Hr HAZWOPER": D(7, 1, 2005),
        "OSHA 8Hr Refresher": D(4, 17, 2025),
        "Drilling Safety": D(6, 4, 2010),
    },
    "Hardy Rodríguez Vázquez": {
        "Protección Contra Caídas": D(4, 17, 2025),
        "Seguridad Eléctrica": D(4, 17, 2025),
        "Lockout": D(4, 17, 2025),
        "Comunicación de Riesgos": D(4, 17, 2025),
        "OSHA 40Hr HAZWOPER": D(1, 1, 2008),
        "OSHA 8Hr Refresher": D(4, 17, 2025),
        "Drilling Safety": D(6, 4, 2010),
        "Utility Locating": D(4, 8, 2025),
    },
    "Juan D. Negrón Hernández": {
        "Protección Contra Caídas": D(4, 17, 2025),
        "Seguridad Eléctrica": D(4, 17, 2025),
        "Lockout": D(4, 17, 2025),
        "Comunicación de Riesgos": D(4, 17, 2025),
        "OSHA 40Hr HAZWOPER": D(11, 1, 1992),
        "OSHA 8Hr Refresher": D(4, 17, 2025),
        "Drilling Safety": D(8, 8, 2008),
        "Utility Locating": D(4, 8, 2025),
    },
    "Rafael Díaz Colón": {
        "Protección Contra Caídas": D(4, 17, 2025),
        "Seguridad Eléctrica": D(4, 17, 2025),
        "Lockout": D(4, 17, 2025),
        "Comunicación de Riesgos": D(4, 17, 2025),
        "OSHA 40Hr HAZWOPER": D(1, 1, 2008),
        "OSHA 8Hr Refresher": D(4, 17, 2025),
        "Drilling Safety": D(6, 4, 2010),
    },
    "Luis A. Padilla Díaz": {
        "Protección Contra Caídas": D(4, 17, 2025),
        "Seguridad Eléctrica": D(4, 17, 2025),
        "Lockout": D(4, 17, 2025),
        "Comunicación de Riesgos": D(4, 17, 2025),
        "OSHA 40Hr HAZWOPER": D(7, 1, 2017),
        "OSHA 8Hr Refresher": D(4, 17, 2025),
        "Drilling Safety": D(1, 10, 2017),
    },
    "Jesús Quiñonez Ayala": {
        "Protección Contra Caídas": D(4, 17, 2025),
        "Seguridad Eléctrica": D(4, 17, 2025),
        "Lockout": D(4, 17, 2025),
        "Comunicación de Riesgos": D(4, 17, 2025),
        "OSHA 40Hr HAZWOPER": D(7, 1, 2017),
        "OSHA 8Hr Refresher": D(4, 17, 2025),
        "Drilling Safety": D(1, 10, 2017),
    },
}


# ---- Styling helpers ----
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True, color="1F3A5F", size=10)
TITLE_FONT = Font(bold=True, size=16, color="1F3A5F")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


# ---- Sheets ----
def build_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    ws["A1"] = "Contractor Certifications Tracker"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    content = [
        "",
        ("How to use this workbook", True),
        "1. Add new contractors on the Contractors tab.",
        "2. Add workers on the Workers tab. Each worker belongs to one contractor (pick from dropdown).",
        "3. Review required certifications on the Certifications tab. Add new ones if HSE adds them.",
        "4. On the Tracker tab, enter the completion date for each worker / certification combo.",
        "5. The Dashboard tab auto-updates: compliance %, missing certs, expiring soon.",
        "",
        ("Color coding on the Tracker", True),
        "• Red cell     = certification missing (empty)",
        "• Yellow cell  = completed, expiring within 60 days",
        "• Orange cell  = expired (past validity based on Certifications tab)",
        "• Green cell   = current",
        "",
        ("Sharing with peers", True),
        "Save this file to OneDrive or SharePoint and turn on co-authoring so peers can edit simultaneously.",
        "Every edit is tracked in Version History (File → Info → Version History).",
        "",
        ("Notes", True),
        "• Validity years in the Certifications tab drive expiration logic. 0 = no expiration.",
        "• Dates must be entered as real dates (mm/dd/yyyy), not text.",
    ]
    row = 2
    for item in content:
        if isinstance(item, tuple):
            ws.cell(row=row, column=1, value=item[0]).font = Font(bold=True, size=12, color="1F3A5F")
        else:
            ws.cell(row=row, column=1, value=item).alignment = Alignment(wrap_text=True)
        row += 1

    autosize(ws, {1: 110})
    ws.sheet_view.showGridLines = False


def build_contractors(wb: Workbook) -> None:
    ws = wb.create_sheet("Contractors")
    headers = ["Contractor Name", "Primary Contact", "Specialty", "Workers (count)", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, 1, len(headers))

    for i, (name, contact, specialty, notes) in enumerate(CONTRACTORS, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=contact)
        ws.cell(row=i, column=3, value=specialty)
        ws.cell(row=i, column=4, value=f'=COUNTIF(Workers!B:B,A{i})')
        ws.cell(row=i, column=5, value=notes)

    # Leave empty rows for future additions
    last_row = 50
    table_ref = f"A1:E{last_row}"
    table = Table(displayName="ContractorsTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    autosize(ws, {1: 32, 2: 28, 3: 28, 4: 16, 5: 40})
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_certifications(wb: Workbook) -> None:
    ws = wb.create_sheet("Certifications")
    headers = ["Certification", "Category", "Validity (years, 0=none)", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, 1, len(headers))

    for i, (name, category, validity) in enumerate(ALL_CERTS, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=category)
        ws.cell(row=i, column=3, value=validity)

    last_row = 60
    table = Table(displayName="CertificationsTable", ref=f"A1:D{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    # Dropdown for category
    dv = DataValidation(
        type="list", formula1='"HSE Required,Additional Training"', allow_blank=True
    )
    dv.add(f"B2:B{last_row}")
    ws.add_data_validation(dv)

    autosize(ws, {1: 38, 2: 22, 3: 22, 4: 38})
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_job_titles(wb: Workbook) -> None:
    ws = wb.create_sheet("Job Titles")
    headers = ["Job Title", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, 1, len(headers))

    ws.cell(row=2, column=1, value="Existing titles are read directly from Workers column C.")
    ws.cell(row=2, column=2, value="Type a new title directly in Workers column C once, and it will appear in the dropdown list.")

    wb.defined_names["JobTitleOptions"] = DefinedName(
        "JobTitleOptions",
        attr_text="Workers!$C$2:INDEX(Workers!$C:$C,LOOKUP(2,1/(Workers!$C:$C<>\"\"),ROW(Workers!$C:$C)))",
    )

    autosize(ws, {1: 28, 2: 80})
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build_workers(wb: Workbook) -> None:
    ws = wb.create_sheet("Workers")
    headers = [
        "Worker Name",
        "Contractor",
        "Job Title",
        "Status",
        "Employee Code",
        "Hire Date",
        "Email",
        "Phone",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, 1, len(headers))

    for i, (name, contractor, title, status) in enumerate(WORKERS, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=contractor)
        ws.cell(row=i, column=3, value=title)
        ws.cell(row=i, column=4, value=status)

    last_row = 200
    table = Table(displayName="WorkersTable", ref=f"A1:I{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    # Contractor dropdown from Contractors!A
    dv_contractor = DataValidation(
        type="list", formula1="=Contractors!$A$2:$A$50", allow_blank=True
    )
    dv_contractor.add(f"B2:B{last_row}")
    ws.add_data_validation(dv_contractor)

    dv_job_title = DataValidation(type="list", formula1="=JobTitleOptions", allow_blank=True)
    dv_job_title.showErrorMessage = False
    dv_job_title.add(f"C2:C{last_row}")
    ws.add_data_validation(dv_job_title)

    dv_status = DataValidation(
        type="list", formula1='"active,inactive,onboarding,terminated"', allow_blank=True
    )
    dv_status.add(f"D2:D{last_row}")
    ws.add_data_validation(dv_status)

    # Date format for hire date
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=6).number_format = "mm/dd/yyyy"

    autosize(ws, {1: 28, 2: 28, 3: 22, 4: 14, 5: 16, 6: 14, 7: 28, 8: 16, 9: 30})
    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False


def build_tracker(wb: Workbook) -> None:
    ws = wb.create_sheet("Tracker")
    cert_names = [c[0] for c in ALL_CERTS]
    hse_count = len(HSE_REQUIRED)

    # Row 1: category banner across cert columns
    ws.cell(row=1, column=1, value="").fill = HEADER_FILL
    ws.cell(row=1, column=2, value="").fill = HEADER_FILL
    ws.cell(row=1, column=3, value="HSE Required Certifications")
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2 + hse_count)
    ws.cell(row=1, column=3 + hse_count, value="Additional Training")
    ws.merge_cells(
        start_row=1,
        start_column=3 + hse_count,
        end_row=1,
        end_column=2 + len(cert_names),
    )
    for col in (3, 3 + hse_count):
        c = ws.cell(row=1, column=col)
        c.fill = PatternFill("solid", fgColor="2E5984")
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: full headers
    headers = ["Contractor", "Worker"] + cert_names
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, len(headers))
    ws.row_dimensions[2].height = 60

    # Worker rows
    for i, (name, contractor, *_rest) in enumerate(WORKERS, start=3):
        ws.cell(row=i, column=1, value=contractor)
        ws.cell(row=i, column=2, value=name)
        data = CERT_DATA.get(name, {})
        for j, cert in enumerate(cert_names, start=3):
            cell = ws.cell(row=i, column=j, value=data.get(cert))
            cell.number_format = "mm/dd/yyyy"
            cell.alignment = Alignment(horizontal="center")

    last_row = 200
    last_col = 2 + len(cert_names)

    # Format the full date matrix area
    for row in range(3, last_row + 1):
        for col in range(3, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = "mm/dd/yyyy"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Data validation: contractor dropdown
    dv_contractor = DataValidation(
        type="list", formula1="=Contractors!$A$2:$A$50", allow_blank=True
    )
    dv_contractor.add(f"A3:A{last_row}")
    ws.add_data_validation(dv_contractor)

    # Data validation: worker dropdown
    dv_worker = DataValidation(
        type="list", formula1="=Workers!$A$2:$A$200", allow_blank=True
    )
    dv_worker.add(f"B3:B{last_row}")
    ws.add_data_validation(dv_worker)

    # Conditional formatting on cert matrix
    matrix_range = f"C3:{get_column_letter(last_col)}{last_row}"

    # Missing (empty cells where worker exists) -> red
    red_fill = PatternFill("solid", fgColor="F8CBAD")
    ws.conditional_formatting.add(
        matrix_range,
        FormulaRule(formula=[f'AND($B3<>"",C3="")'], fill=red_fill, stopIfTrue=False),
    )

    # Expired -> orange (cert older than validity years from Certifications sheet)
    orange_fill = PatternFill("solid", fgColor="F4B183")
    expired_formula = (
        'AND(C3<>"",'
        'ISNUMBER(C3),'
        'VLOOKUP(C$2,Certifications!$A:$C,3,FALSE)>0,'
        'EDATE(C3,VLOOKUP(C$2,Certifications!$A:$C,3,FALSE)*12)<TODAY())'
    )
    ws.conditional_formatting.add(
        matrix_range, FormulaRule(formula=[expired_formula], fill=orange_fill, stopIfTrue=False)
    )

    # Expiring within 60 days -> yellow
    yellow_fill = PatternFill("solid", fgColor="FFE699")
    expiring_formula = (
        'AND(C3<>"",'
        'ISNUMBER(C3),'
        'VLOOKUP(C$2,Certifications!$A:$C,3,FALSE)>0,'
        'EDATE(C3,VLOOKUP(C$2,Certifications!$A:$C,3,FALSE)*12)>=TODAY(),'
        'EDATE(C3,VLOOKUP(C$2,Certifications!$A:$C,3,FALSE)*12)<=TODAY()+60)'
    )
    ws.conditional_formatting.add(
        matrix_range, FormulaRule(formula=[expiring_formula], fill=yellow_fill, stopIfTrue=False)
    )

    # Current (has date) -> green
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    current_formula = 'AND(C3<>"",ISNUMBER(C3))'
    ws.conditional_formatting.add(
        matrix_range, FormulaRule(formula=[current_formula], fill=green_fill, stopIfTrue=False)
    )

    # Column widths
    widths = {1: 26, 2: 26}
    for idx in range(3, last_col + 1):
        widths[idx] = 14
    autosize(ws, widths)

    tracker_table = Table(
        displayName="TrackerTable",
        ref=f"A2:{get_column_letter(last_col)}{last_row}",
    )
    tracker_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(tracker_table)

    ws.freeze_panes = "C3"
    ws.sheet_view.showGridLines = False


def build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard", 0)  # first sheet
    ws["A1"] = "Contractor Certifications Dashboard"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    # KPI cards
    kpi_row = 3
    labels = ["Total Contractors", "Total Workers", "Active Workers", "Today"]
    formulas = [
        "=COUNTA(Contractors!A2:A50)",
        "=COUNTA(Workers!A2:A200)",
        '=COUNTIFS(Workers!A2:A200,"<>",Workers!D2:D200,"active")',
        "=TEXT(TODAY(),\"mm/dd/yyyy\")",
    ]
    for i, (label, formula) in enumerate(zip(labels, formulas)):
        col = 1 + i * 2
        lbl = ws.cell(row=kpi_row, column=col, value=label)
        lbl.font = SUBHEADER_FONT
        lbl.fill = SUBHEADER_FILL
        lbl.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col + 1)

        val = ws.cell(row=kpi_row + 1, column=col, value=formula)
        val.font = Font(bold=True, size=20, color="1F3A5F")
        val.alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 2, end_column=col + 1
        )

    # Compliance per certification
    header_row = 7
    ws.cell(row=header_row, column=1, value="Compliance by Certification")
    ws.cell(row=header_row, column=1).font = Font(bold=True, size=13, color="1F3A5F")
    ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=6)

    sub_headers = ["Certification", "Category", "Completed", "Missing", "% Complete", "Bar"]
    for col, h in enumerate(sub_headers, 1):
        ws.cell(row=header_row + 1, column=col, value=h)
    style_header_row(ws, header_row + 1, 1, len(sub_headers))

    # Workers count dynamic
    for i, (name, category, _validity) in enumerate(ALL_CERTS):
        r = header_row + 2 + i
        # Column letter in Tracker
        tracker_col_letter = get_column_letter(3 + i)
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=category)
        # Completed = non-empty cells in that column (excluding header)
        ws.cell(
            row=r,
            column=3,
            value=f'=COUNTA(Tracker!{tracker_col_letter}3:{tracker_col_letter}200)',
        )
        # Missing = workers with blank in that column
        ws.cell(
            row=r,
            column=4,
            value=f'=COUNTA(Workers!A2:A200)-C{r}',
        )
        # Percent
        pct = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/(C{r}+D{r}),0)")
        pct.number_format = "0.0%"
        # Text bar using REPT
        ws.cell(row=r, column=6, value=f'=REPT("■",ROUND(E{r}*20,0))')
        ws.cell(row=r, column=6).font = Font(color="2E7D32", size=11)
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = THIN_BORDER

    last_r = header_row + 1 + len(ALL_CERTS)
    # Conditional formatting on % column
    pct_range = f"E{header_row + 2}:E{last_r}"
    ws.conditional_formatting.add(
        pct_range,
        CellIsRule(operator="lessThan", formula=["0.7"], fill=PatternFill("solid", fgColor="F8CBAD")),
    )
    ws.conditional_formatting.add(
        pct_range,
        CellIsRule(
            operator="between",
            formula=["0.7", "0.9"],
            fill=PatternFill("solid", fgColor="FFE699"),
        ),
    )
    ws.conditional_formatting.add(
        pct_range,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["0.9"],
            fill=PatternFill("solid", fgColor="C6EFCE"),
        ),
    )

    autosize(ws, {1: 36, 2: 22, 3: 14, 4: 14, 5: 14, 6: 28})
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = Workbook()
    # Remove default sheet, we'll create our own
    wb.remove(wb.active)

    build_instructions(wb)
    build_contractors(wb)
    build_certifications(wb)
    build_job_titles(wb)
    build_workers(wb)
    build_tracker(wb)
    build_dashboard(wb)  # inserted at position 0

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
