"""Import a contractor certification PDF (Anejo 3 format) into the Excel tracker.

Usage:
    python tools/import_pdf.py "path/to/file.pdf" [more.pdf ...]

Or drag a PDF onto Import PDF.bat in the project root.

Rules:
- Contractor name is read from "Nombre de la compañía contratista: X".
- Workers are matched by normalized name (accent/case insensitive) within a contractor.
- New workers are appended; existing workers keep the most recent date per cert.
- New contractors are added to the Contractors sheet automatically.
"""
from __future__ import annotations

import os
import re
import sys
import unicodedata
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import pdfplumber
from openpyxl.formatting.rule import FormulaRule
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "Contractor Certifications Tracker.xlsx"
# WORKBOOK_PATH = ROOT / "Contractor Certifications Tracker Demo.xlsx"


# --- Normalization ---
def normalize(s: Optional[str]) -> str:
    if not s:
        return ""
    s = s.strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_company(s: Optional[str]) -> str:
    """Match contractor names tolerantly: strip trailing punctuation and
    collapse common business-suffix variations ("Inc." vs "Inc")."""
    base = normalize(s)
    if not base:
        return ""
    # Drop periods and commas everywhere — they're decorative in company names.
    base = base.replace(".", "").replace(",", "")
    base = re.sub(r"\s+", " ", base).strip()
    return base


def normalize_compact(s: Optional[str]) -> str:
    """Whitespace-insensitive normalization (e.g. 'OSHA 8 Hr' == 'OSHA 8Hr')."""
    return re.sub(r"\s+", "", normalize(s))


# --- Date parsing ---
MONTHS_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

EXTRA_ADDITIONAL_CERTS = [
    ("OSHA 10", 0),
    ("OSHA 30", 0),
    ("Rebar Safety", 0),
    ("Formwork & Shoring", 0),
    ("Silica Exposure", 0),
    ("Concrete & Masonry", 0),
    ("Equipment Training", 0),
    ("OSHA 30 501", 0),
    ("OSHA 30 511", 0),
]

REQUIRED_BASELINE_CERTS = [
    ("Protección Contra Caídas", "HSE Required", 1),
    ("Flama Expuesta", "HSE Required", 1),
    ("Espacios Confinados", "HSE Required", 1),
    ("Manejo de Equipo Motorizado", "HSE Required", 3),
    ("Excavación o Zanja", "HSE Required", 1),
    ("Manejo de Tijeras (Scissor Lift)", "HSE Required", 3),
    ("Seguridad Eléctrica", "HSE Required", 1),
    ("Lockout", "HSE Required", 1),
    ("Manejo de Grúas", "HSE Required", 3),
    ("Escaleras", "HSE Required", 3),
    ("Andamios", "HSE Required", 3),
    ("Trabajos con Plomo o Asbesto", "HSE Required", 1),
    ("Comunicación de Riesgos", "HSE Required", 1),
    ("Inducción Jacobs/Lilly", "Additional Training", 1),
    ("OSHA 40Hr HAZWOPER", "Additional Training", 0),
    ("OSHA 8Hr Refresher", "Additional Training", 1),
    ("Drilling Safety", "Additional Training", 3),
    ("Utility Locating", "Additional Training", 1),
    *[(name, "Additional Training", validity) for name, validity in EXTRA_ADDITIONAL_CERTS],
]

PAGE2_HEADER_ALIASES = {
    "induccion jacobs/lilly": "Inducción Jacobs/Lilly",
    "induccion jacobs lilly": "Inducción Jacobs/Lilly",
    "osha10": "OSHA 10",
    "osha 10": "OSHA 10",
    "osha30": "OSHA 30",
    "osha 30": "OSHA 30",
    "rebar safety": "Rebar Safety",
    "formwork & shoring": "Formwork & Shoring",
    "formwork and shoring": "Formwork & Shoring",
    "formwork shoring": "Formwork & Shoring",
    "silica exposure": "Silica Exposure",
    "concrete & masonry": "Concrete & Masonry",
    "concrete and masonry": "Concrete & Masonry",
    "concrete masonry": "Concrete & Masonry",
    "concrete & mansory": "Concrete & Masonry",
    "concrete mansory": "Concrete & Masonry",
    "equipment training": "Equipment Training",
    "osha 30 501": "OSHA 30 501",
    "osha30501": "OSHA 30 501",
    "osha 30-501": "OSHA 30 501",
    "osha 30 511": "OSHA 30 511",
    "osha30511": "OSHA 30 511",
    "osha 30-511": "OSHA 30 511",
    # OSHA 40Hr HAZWOPER often comes through with a misread / typo
    # ("HOZWOPER") and varying spacing because the page-2 header is rotated.
    "osha 40hr hazwoper": "OSHA 40Hr HAZWOPER",
    "osha 40 hr hazwoper": "OSHA 40Hr HAZWOPER",
    "osha 40hr hozwoper": "OSHA 40Hr HAZWOPER",
    "osha 40 hr hozwoper": "OSHA 40Hr HAZWOPER",
    # OSHA 8Hr Refresher with optional space between "8" and "Hr".
    "osha 8hr refresher": "OSHA 8Hr Refresher",
    "osha 8 hr refresher": "OSHA 8Hr Refresher",
    "drilling safety": "Drilling Safety",
    "utility locating": "Utility Locating",
}

# Matches either a plain "m/d/yyyy" date or a day-range form like
# "2/5-9/2023" or "6/5-8-2023" where we want the last day of the range.
DATE_RE = re.compile(r"\d{1,2}/\d{1,2}(?:-\d{1,2})?[-/]\d{2,4}")
DATE_RANGE_RE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})-(\d{1,2})[-/](\d{2,4})\s*$")


def parse_date(raw: Optional[str]) -> Optional[date]:
    if not raw:
        return None
    s = raw.strip()
    if not s:
        return None
    s = s.replace("\n", " ").strip()

    # Collapse day-range forms ("2/5-9/2023", "6/5-8-2023") to a single date by
    # taking the LAST day of the range. Then let the standard parsing handle it.
    range_match = DATE_RANGE_RE.match(s)
    if range_match:
        month, _start_day, end_day, year = range_match.groups()
        s = f"{month}/{end_day}/{year}"

    slash_compact = re.sub(r"\s+", "", s) if "/" in s else None

    candidates = [s]
    if slash_compact and slash_compact != s:
        candidates.append(slash_compact)

    for candidate in candidates:
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                pass

    for fmt in ("%B %Y", "%b %Y", "%B, %Y", "%b. %Y"):
        try:
            return datetime.strptime(s, fmt).date().replace(day=1)
        except ValueError:
            pass

    # Spanish "Enero 2008" / "enero de 2008"
    m = re.match(r"([a-záéíóúñ]+)[\s,]+(?:de\s+)?(\d{4})", s.lower())
    if m and m.group(1) in MONTHS_ES:
        return date(int(m.group(2)), MONTHS_ES[m.group(1)], 1)

    return None


# --- Cert name matching ---
def build_cert_alias_map(ws: Worksheet) -> dict[str, str]:
    """Map normalized cert header -> canonical cert name from Certifications sheet."""
    aliases = {}
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        name = row[0]
        if not name:
            continue
        aliases[normalize(name)] = name
        aliases[normalize_compact(name)] = name
        # Add a few common variants
        compact = normalize(re.sub(r"[()]", "", name))
        aliases[compact] = name
        aliases[normalize_compact(compact)] = name
    # Known alternate spellings
    extras = {
        "osha 40hr hozwoper": "OSHA 40Hr HAZWOPER",
        "osha 40 hr hazwoper": "OSHA 40Hr HAZWOPER",
        "osha40hr hazwoper": "OSHA 40Hr HAZWOPER",
        "osha 8hr refresher": "OSHA 8Hr Refresher",
        "osha8hr refresher": "OSHA 8Hr Refresher",
        "osha 8 hr refresher": "OSHA 8Hr Refresher",
        "osha8hrefresher": "OSHA 8Hr Refresher",
        "flama expuesta": "Flama Expuesta",
        "flama": "Flama Expuesta",
        "flama exp": "Flama Expuesta",
        "llama expuesta": "Flama Expuesta",
        "manejo de tijeras scissor lift": "Manejo de Tijeras (Scissor Lift)",
        "manejo de tijeras": "Manejo de Tijeras (Scissor Lift)",
        "scissor lift": "Manejo de Tijeras (Scissor Lift)",
        "excavacion o zanja": "Excavación o Zanja",
        "excavacion": "Excavación o Zanja",
        "trabajos con plomo o asbesto": "Trabajos con Plomo o Asbesto",
        "trabajo con plomo o asbesto": "Trabajos con Plomo o Asbesto",
        "plomo o asbesto": "Trabajos con Plomo o Asbesto",
        "plomo asbesto": "Trabajos con Plomo o Asbesto",
        "lead asbestos": "Trabajos con Plomo o Asbesto",
        "induccion jacobs lilly": "Inducción Jacobs/Lilly",
        "induccion jacobs/lilly": "Inducción Jacobs/Lilly",
        "induccion": "Inducción Jacobs/Lilly",
        "comunicacion de riesgos": "Comunicación de Riesgos",
        "proteccion contra caidas": "Protección Contra Caídas",
        "seguridad electrica": "Seguridad Eléctrica",
        "drilling safety": "Drilling Safety",
        "utility locating": "Utility Locating",
        "osha 10": "OSHA 10",
        "osha10": "OSHA 10",
        "osha 30": "OSHA 30",
        "osha30": "OSHA 30",
        "rebar safety": "Rebar Safety",
        "formwork & shoring": "Formwork & Shoring",
        "formwork and shoring": "Formwork & Shoring",
        "formwork shoring": "Formwork & Shoring",
        "silica exposure": "Silica Exposure",
        "concrete & masonry": "Concrete & Masonry",
        "concrete and masonry": "Concrete & Masonry",
        "concrete masonry": "Concrete & Masonry",
        "concrete & mansory": "Concrete & Masonry",
        "concrete mansory": "Concrete & Masonry",
        "equipment training": "Equipment Training",
        "osha 30 501": "OSHA 30 501",
        "osha30501": "OSHA 30 501",
        "osha 30-501": "OSHA 30 501",
        "osha 30 511": "OSHA 30 511",
        "osha30511": "OSHA 30 511",
        "osha 30-511": "OSHA 30 511",
    }
    for k, v in extras.items():
        aliases.setdefault(k, v)
    return aliases


STOPWORDS = {"de", "la", "el", "o", "y", "con", "por", "del", "las", "los", "a"}


def match_cert(header: str, aliases: dict[str, str]) -> Optional[str]:
    """Map a raw PDF header to a canonical cert name with fuzzy fallback."""
    key = normalize(header)
    if not key:
        return None
    no_ws = normalize_compact(header)
    if no_ws in aliases:
        return aliases[no_ws]
    if key in aliases:
        return aliases[key]
    # Remove parentheses and try again
    compact = normalize(re.sub(r"[()]", "", key))
    if compact in aliases:
        return aliases[compact]
    # Whitespace-insensitive: "osha 8 hr refresher" -> "osha8hrrefresher"
    for alias_key, canonical in aliases.items():
        if re.sub(r"\s+", "", alias_key) == no_ws:
            return canonical
    # Substring match (either direction)
    for norm_key, canonical in aliases.items():
        if len(norm_key) < 6:
            continue
        if norm_key in key or key in norm_key:
            return canonical
    # Token overlap
    key_tokens = set(key.split()) - STOPWORDS
    if not key_tokens:
        return None
    best_canonical = None
    best_score = 0
    for norm_key, canonical in aliases.items():
        tokens = set(norm_key.split()) - STOPWORDS
        if len(tokens) < 2:
            continue
        overlap = len(key_tokens & tokens)
        if overlap >= 2 and overlap > best_score:
            best_score = overlap
            best_canonical = canonical
    return best_canonical


# --- PDF extraction ---
def find_first_data_row(table: list[list], header_idx: int, name_col: int) -> int:
    for i in range(header_idx + 1, len(table)):
        row = table[i] or []
        if name_col >= len(row):
            continue
        worker_name = " ".join(str(row[name_col] or "").split()).strip()
        if worker_name and "nombre" not in normalize(worker_name):
            return i
    return min(header_idx + 1, len(table))


def build_table_header(table: list[list], header_idx: int, first_data_idx: int) -> list[str]:
    header_rows = table[header_idx:first_data_idx]
    width = max((len(row or []) for row in header_rows), default=0)
    combined: list[str] = []

    for col_idx in range(width):
        parts: list[str] = []
        for row in header_rows:
            if col_idx >= len(row):
                continue
            cell = " ".join(str(row[col_idx] or "").split()).strip()
            if not cell:
                continue
            if not parts or parts[-1] != cell:
                parts.append(cell)
        combined.append(" ".join(parts).strip())

    return combined


def value_for_cert_column(
    row: list,
    idx: int,
    header_name: str,
    header_list: Optional[list] = None,
) -> Optional[str]:
    if idx < len(row) and row[idx] not in (None, ""):
        return row[idx]

    # Rescue case: when a multi-line header wraps awkwardly, pdfplumber sometimes
    # creates an extra "orphan" column next to the real one — date data lands in
    # the orphan, while the real column ends up with None. Pull from an
    # immediately adjacent column ONLY if that neighbor has no header of its
    # own (so we can't accidentally steal a date from a different cert).
    if header_list is None:
        return None
    for offset in (-1, 1):
        nb = idx + offset
        if nb < 0 or nb >= len(row):
            continue
        nb_header = header_list[nb] if nb < len(header_list) else ""
        if normalize(str(nb_header) if nb_header is not None else ""):
            continue  # neighbor has its own header — not an orphan
        candidate = row[nb] if nb < len(row) else None
        if candidate and parse_date(candidate):
            return candidate
    return None


def extract_contractor(text: str) -> Optional[str]:
    # Work on an accent-stripped copy for matching, but recover the original slice
    # so we keep accents in the final contractor name.
    ascii_text = unicodedata.normalize("NFKD", text)
    ascii_text = "".join(c for c in ascii_text if not unicodedata.combining(c))

    patterns = [
        r"nombre\s+de\s+la\s+compa\S*\s+contratista\s*[:\-]?\s*([^\n*]+?)(?:\s*\*|\s*fecha|\s*pagina|\n|$)",
        r"compa\S*\s+contratista\s*[:\-]?\s*([^\n*]+?)(?:\s*\*|\s*fecha|\s*pagina|\n|$)",
        r"contratista\s*[:\-]\s*([^\n*]+?)(?:\s*\*|\s*fecha|\s*pagina|\n|$)",
        r"contractor\s*(?:name)?\s*[:\-]\s*([^\n*]+?)(?:\s*\*|\s*date|\n|$)",
    ]
    for pattern in patterns:
        m = re.search(pattern, ascii_text, flags=re.IGNORECASE)
        if m:
            start, end = m.span(1)
            # Pull the corresponding slice from the original text (indexes line up
            # because NFKD decomposition only *adds* combining marks; we stripped them)
            raw = text[start:end] if end <= len(text) else m.group(1)
            # If lengths drifted because of normalization edge cases, fall back
            if len(raw) < 2 or len(raw) > len(m.group(1)) + 10:
                raw = m.group(1)
            return raw.strip().rstrip(".,;")
    return None


def cleanup_primary_contact(raw: str) -> Optional[str]:
    """Rebuild a contact name when pdfplumber scatters letters across signature lines."""
    raw = re.split(r"\s{3,}|\t", raw)[0]
    raw = raw.replace("\r", " ").replace("\n", " ")
    raw = raw.strip().strip("_").strip().rstrip(".,;:")
    if not raw:
        return None

    if "_" not in raw:
        parts = raw.split()
        while parts and len(parts[-1]) <= 2 and parts[-1].islower():
            parts.pop()
        simple_name = " ".join(parts).strip()
        return simple_name or None

    fragments = list(re.finditer(r"[^\W\d_]+", raw, flags=re.UNICODE))
    if not fragments:
        return None

    words: list[str] = []
    current = fragments[0].group(0)
    prev_fragment = fragments[0].group(0)
    prev_end = fragments[0].end()

    for match in fragments[1:]:
        fragment = match.group(0)
        separator = raw[prev_end:match.start()]
        separator = separator.replace("\u00A0", " ")
        weak_join = False
        if " " in separator and len(current) >= 3 and fragment[:1].isupper():
            weak_join = False
        elif re.fullmatch(r"_+", separator):
            weak_join = True
        elif re.fullmatch(r"[ _]{1,2}", separator):
            weak_join = "_" in separator or (
                len(prev_fragment) <= 2 and len(fragment) <= 2
            )

        if weak_join:
            current += fragment
        else:
            words.append(current)
            current = fragment

        prev_fragment = fragment
        prev_end = match.end()

    words.append(current)

    cleaned_words = []
    for word in words:
        letters_only = re.sub(r"[^\w'-]", "", word, flags=re.UNICODE)
        letters_only = re.sub(r"[\d_]", "", letters_only)
        if not letters_only:
            continue
        if letters_only.islower() or letters_only.isupper():
            letters_only = letters_only.capitalize()
        cleaned_words.append(letters_only)

    if not cleaned_words or len(cleaned_words) > 6:
        return None

    return " ".join(cleaned_words)


def extract_primary_contact(text: str) -> Optional[str]:
    """Pull the name after 'Certificado por nombre/firma:'."""
    ascii_text = unicodedata.normalize("NFKD", text)
    ascii_text = "".join(c for c in ascii_text if not unicodedata.combining(c))
    patterns = [
        r"certificado\s+por\s+nombre\s*/?\s*firma\s*[:\-_]*\s*([^\n]+)",
        r"nombre\s*/?\s*firma\s*[:\-_]*\s*([^\n]+)",
    ]
    for pattern in patterns:
        m = re.search(pattern, ascii_text, flags=re.IGNORECASE)
        if not m:
            continue
        start, end = m.span(1)
        raw = text[start:end] if end <= len(text) else m.group(1)
        name = cleanup_primary_contact(raw)
        if not name:
            continue
        if 2 <= len(name) <= 80:
            return name
    return None


def merge_worker_certs(
    workers: dict[str, dict[str, date]],
    parsed: dict[str, dict[str, date]],
) -> None:
    for worker_name, certs in parsed.items():
        entry = workers.setdefault(worker_name, {})
        for cert_name, dt in certs.items():
            existing = entry.get(cert_name)
            if existing is None or dt > existing:
                entry[cert_name] = dt


def canonicalize_page2_header(label: str) -> Optional[str]:
    label = label.replace("\n", " ")
    label = label.replace("&", " & ")
    # Page-2 headers come from rotated text; words like "OSHA40Hr" arrive
    # without a space. Split letter/digit boundaries so "OSHA40Hr HOZWOPER"
    # normalizes to "osha 40 hr hozwoper" before alias lookup.
    label = re.sub(r"([A-Za-z])(\d)", r"\1 \2", label)
    label = re.sub(r"(\d)([A-Za-z])", r"\1 \2", label)
    label = re.sub(r"\s+", " ", label).strip()
    key = normalize(label)
    return PAGE2_HEADER_ALIASES.get(key)


def extract_additional_training_page(page: pdfplumber.page.Page) -> dict[str, dict[str, date]]:
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
    if not words:
        return {}

    header_words = [
        w for w in words
        if 150 <= w["top"] <= 205 and w["x0"] >= 180 and not DATE_RE.fullmatch(w["text"])
    ]
    if not header_words:
        return {}

    x_groups: list[dict[str, object]] = []
    for word in sorted(header_words, key=lambda w: (w["x0"], w["top"])):
        if not x_groups or abs(word["x0"] - x_groups[-1]["x0"]) > 4:
            x_groups.append({"x0": word["x0"], "words": [word]})
        else:
            x_groups[-1]["words"].append(word)

    header_parts = []
    for group in x_groups:
        group_words = group["words"]
        label = "".join(w["text"] for w in sorted(group_words, key=lambda w: w["top"], reverse=True))
        header_parts.append(
            {
                "label": label,
                "x0": min(w["x0"] for w in group_words),
                "x1": max(w["x1"] for w in group_words),
            }
        )

    training_columns = []
    for part in header_parts:
        if not training_columns or part["x0"] - training_columns[-1]["x1"] > 18:
            training_columns.append({"parts": [part], "x0": part["x0"], "x1": part["x1"]})
        else:
            training_columns[-1]["parts"].append(part)
            training_columns[-1]["x1"] = part["x1"]

    column_defs = []
    for group in training_columns:
        raw_label = " ".join(part["label"] for part in group["parts"])
        canonical = canonicalize_page2_header(raw_label)
        if not canonical:
            # Unknown page-2 training. Keep it anyway so import_pdf can
            # auto-register it as a new certification.
            cleaned = raw_label.replace("\n", " ")
            cleaned = re.sub(r"([A-Za-z])(\d)", r"\1 \2", cleaned)
            cleaned = re.sub(r"(\d)([A-Za-z])", r"\1 \2", cleaned)
            cleaned = re.sub(r"\s+", " ", cleaned).strip()
            if not cleaned:
                continue
            canonical = cleaned
        column_defs.append(
            {
                "name": canonical,
                "center_x": (group["x0"] + group["x1"]) / 2,
            }
        )

    if DEBUG:
        print(f"  [debug] page 2 additional columns: {[c['name'] for c in column_defs]}")

    if not column_defs:
        return {}

    name_words = [
        w for w in words
        if 205 <= w["top"] <= 430 and w["x0"] < 180 and not DATE_RE.fullmatch(w["text"])
    ]
    if not name_words:
        return {}

    name_rows = []
    for word in sorted(name_words, key=lambda w: (w["top"], w["x0"])):
        if not name_rows or abs(word["top"] - name_rows[-1]["top"]) > 8:
            name_rows.append({"top": word["top"], "words": [word]})
        else:
            name_rows[-1]["words"].append(word)

    worker_rows = []
    for row in name_rows:
        row_words = sorted(row["words"], key=lambda w: w["x0"])
        worker_name = " ".join(w["text"] for w in row_words).strip()
        if not worker_name or "nombre del empleado" in normalize(worker_name):
            continue
        mids = [((w["top"] + w["bottom"]) / 2) for w in row_words]
        worker_rows.append(
            {
                "name": worker_name,
                "center_y": sum(mids) / len(mids),
            }
        )

    if not worker_rows:
        return {}

    parsed: dict[str, dict[str, date]] = {}

    # Build the list of "date tokens" found in the data area. Two flavors:
    #   1) a single-word m/d/yyyy match (existing behavior)
    #   2) a "MonthName YYYY" pair where the year is the same line OR the line
    #      directly below at similar x (e.g. "January\n2008" — common in OSHA
    #      40Hr HAZWOPER cells when month-year wraps inside a narrow column).
    data_words = sorted(
        [w for w in words if 205 <= w["top"] <= 430 and w["x0"] >= 180],
        key=lambda w: (w["top"], w["x0"]),
    )

    date_tokens: list[dict] = []
    consumed: set[int] = set()

    # Pass 1: single-word numeric dates.
    for w in data_words:
        if DATE_RE.fullmatch(w["text"]) and parse_date(w["text"]):
            date_tokens.append({
                "text": w["text"],
                "x0": w["x0"], "x1": w["x1"],
                "top": w["top"], "bottom": w["bottom"],
            })
            consumed.add(id(w))

    # Pass 2: month-name + 4-digit-year pairs (same line or stacked).
    MONTH_NAMES = {
        "january", "february", "march", "april", "may", "june", "july",
        "august", "september", "october", "november", "december",
        "jan", "feb", "mar", "apr", "jun", "jul", "aug", "sep", "sept",
        "oct", "nov", "dec",
        "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
        "agosto", "septiembre", "setiembre", "octubre", "noviembre",
        "diciembre",
    }
    year_re = re.compile(r"\d{4}")
    for mw in data_words:
        if id(mw) in consumed:
            continue
        if mw["text"].lower().rstrip(".,") not in MONTH_NAMES:
            continue
        mw_cx = (mw["x0"] + mw["x1"]) / 2
        best_year = None
        best_score = float("inf")
        for yw in data_words:
            if id(yw) in consumed:
                continue
            if not year_re.fullmatch(yw["text"]):
                continue
            yw_cx = (yw["x0"] + yw["x1"]) / 2
            dx = abs(mw_cx - yw_cx)
            if dx > 45:
                continue
            dy = yw["top"] - mw["top"]
            if dy < -3 or dy > 25:
                continue
            score = dx + abs(dy) * 0.5
            if score < best_score:
                best_score = score
                best_year = yw
        if best_year is None:
            continue
        combined = f"{mw['text']} {best_year['text']}"
        if not parse_date(combined):
            continue
        date_tokens.append({
            "text": combined,
            "x0": min(mw["x0"], best_year["x0"]),
            "x1": max(mw["x1"], best_year["x1"]),
            "top": mw["top"],
            "bottom": max(mw["bottom"], best_year["bottom"]),
        })
        consumed.add(id(mw))
        consumed.add(id(best_year))

    for token in date_tokens:
        center_y = (token["top"] + token["bottom"]) / 2
        worker_row = min(worker_rows, key=lambda row: abs(row["center_y"] - center_y))
        if abs(worker_row["center_y"] - center_y) > 20:
            continue

        center_x = (token["x0"] + token["x1"]) / 2
        training_col = min(column_defs, key=lambda col: abs(col["center_x"] - center_x))
        dt = parse_date(token["text"])
        if not dt:
            continue

        parsed.setdefault(worker_row["name"], {})[training_col["name"]] = dt

    return parsed


DEBUG = bool(os.environ.get("DEBUG_IMPORT"))


def _fill_missing_cells_from_words(table_obj, data: list[list], words: list) -> None:
    """For each None/empty cell in `data`, search the cell's bbox for words
    that parse as a date and fill the cell with that text.

    Mutates `data` in place. This catches the common pdfplumber failure where
    a cell value (e.g. a date) is on the page but pdfplumber didn't attach it
    to the right cell because surrounding text wraps awkwardly.
    """
    if not table_obj or not words:
        return
    rows = getattr(table_obj, "rows", None) or []
    for row_idx, row_obj in enumerate(rows):
        if row_idx >= len(data):
            continue
        row_data = data[row_idx]
        if row_data is None:
            continue
        cells = getattr(row_obj, "cells", None) or []
        for col_idx, cell_bbox in enumerate(cells):
            if col_idx >= len(row_data):
                continue
            existing = row_data[col_idx]
            if existing not in (None, ""):
                continue
            if not cell_bbox:
                continue
            try:
                x0, top, x1, bottom = cell_bbox
            except (TypeError, ValueError):
                continue
            cell_words = [
                w for w in words
                if w["x0"] >= x0 - 1 and w["x1"] <= x1 + 1
                and w["top"] >= top - 1 and w["bottom"] <= bottom + 1
            ]
            if not cell_words:
                continue
            cell_words.sort(key=lambda w: (w["top"], w["x0"]))
            text = " ".join(w["text"] for w in cell_words).strip()
            if text and parse_date(text):
                row_data[col_idx] = text


def extract_pdf_data(
    pdf_path: Path,
) -> tuple[Optional[str], Optional[str], dict[str, dict[str, date]]]:
    """Return (contractor_name, primary_contact, {worker_name: {cert_name: date}})."""
    contractor: Optional[str] = None
    primary_contact: Optional[str] = None
    workers: dict[str, dict[str, date]] = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            if contractor is None:
                contractor = extract_contractor(text)
            if primary_contact is None:
                primary_contact = extract_primary_contact(text)

            if "nombre del adiestramiento" in normalize(text):
                merge_worker_certs(workers, extract_additional_training_page(page))
                continue

            # Try multiple extraction strategies - default first, then line-based.
            # find_tables returns Table objects (not raw data) so we can pull
            # per-cell bounding boxes and rescue any cell pdfplumber dropped.
            strategies = [
                None,  # default
                {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
                {"vertical_strategy": "text", "horizontal_strategy": "lines"},
            ]
            table_pairs: list[tuple[object, list[list]]] = []
            for strat in strategies:
                objs = page.find_tables(table_settings=strat) if strat else page.find_tables()
                if not objs:
                    continue
                pairs = [(t, t.extract()) for t in objs]
                if any(
                    any("empleado" in normalize((c or "")) for c in (row or []))
                    for _, data in pairs
                    for row in (data or [])
                ):
                    table_pairs = pairs
                    break

            page_words = None
            for tbl_idx, (table_obj, table) in enumerate(table_pairs):
                if not table or len(table) < 2:
                    continue
                # Rescue any cells pdfplumber returned as None by scanning the
                # cell's bbox for date words. Common when cell values wrap
                # across visual lines (e.g. "3/10/202\n6") and confuse the
                # default table extractor.
                if page_words is None:
                    try:
                        page_words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
                    except Exception:
                        page_words = []
                _fill_missing_cells_from_words(table_obj, table, page_words)

                # Find header row: the row that contains "empleado"
                header_idx = None
                for i, row in enumerate(table):
                    if not row:
                        continue
                    if any("empleado" in normalize((c or "")) for c in row):
                        header_idx = i
                        break
                if header_idx is None:
                    continue

                header_row = [(c or "").strip() for c in table[header_idx]]
                name_col = next(
                    (i for i, h in enumerate(header_row) if "empleado" in normalize(h)),
                    0,
                )
                first_data_idx = find_first_data_row(table, header_idx, name_col)
                header = build_table_header(table, header_idx, first_data_idx)

                cert_cols: list[tuple[int, str]] = []
                for idx, h in enumerate(header):
                    if idx == name_col:
                        continue
                    h_norm = normalize(h)
                    if not h_norm:
                        continue
                    if "nombre" in h_norm or h_norm == "adiestramiento":
                        continue
                    cert_cols.append((idx, h))

                if DEBUG:
                    print(f"  [debug] page {page_num} table {tbl_idx}: header_idx={header_idx}")
                    print(f"  [debug]   name_col={name_col}, cert_cols={[(i, h[:30]) for i, h in cert_cols]}")

                for row in table[first_data_idx:]:
                    if not row or name_col >= len(row) or not row[name_col]:
                        continue
                    worker_name = " ".join((row[name_col] or "").split()).strip()
                    if not worker_name or "nombre" in normalize(worker_name):
                        continue
                    entry = workers.setdefault(worker_name, {})
                    for idx, header_name in cert_cols:
                        dt = parse_date(value_for_cert_column(row, idx, header_name, header))
                        if dt:
                            entry[header_name] = dt

    return contractor, primary_contact, workers


# --- Workbook helpers ---
def first_empty_row(ws: Worksheet, key_col: int = 1, start: int = 2, limit: int = 2000) -> int:
    for r in range(start, limit):
        if ws.cell(row=r, column=key_col).value in (None, ""):
            return r
    return limit


def find_row(ws: Worksheet, key_col: int, target_norm: str, start: int = 2, limit: int = 2000) -> Optional[int]:
    for r in range(start, limit):
        val = ws.cell(row=r, column=key_col).value
        if val in (None, ""):
            return None
        if normalize(str(val)) == target_norm:
            return r
    return None


CERT_FORMULA_RE = re.compile(r"^\s*=\s*Certifications!\s*\$?A\$?(\d+)\s*$", re.IGNORECASE)


def tracker_cert_columns(ws: Worksheet) -> dict[str, int]:
    """Map cert name (as in Tracker header row 2) -> column index.

    Handles both literal text headers and formula headers like
    =Certifications!A15 by resolving the formula back to the cert name.
    """
    mapping: dict[str, int] = {}
    ws_certs = ws.parent["Certifications"] if "Certifications" in ws.parent.sheetnames else None

    for col in range(3, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val in (None, ""):
            continue
        if ws_certs is not None and isinstance(val, str):
            m = CERT_FORMULA_RE.match(val)
            if m:
                certs_row = int(m.group(1))
                resolved = ws_certs.cell(row=certs_row, column=1).value
                if resolved:
                    mapping[str(resolved)] = col
                    continue
        mapping[str(val)] = col
    return mapping


def tracker_last_header_col(ws: Worksheet) -> int:
    """Return the last Tracker column that has a real row-2 header."""
    for col in range(ws.max_column, 2, -1):
        if ws.cell(row=2, column=col).value not in (None, ""):
            return col
    return 2


def tracker_styled_template_col(ws: Worksheet) -> int:
    """Find a real styled Tracker cert header to copy for future columns."""
    for col in range(tracker_last_header_col(ws), 2, -1):
        cell = ws.cell(row=2, column=col)
        if cell.value not in (None, "") and cell.fill.fill_type:
            return col
    return 3


def refresh_additional_training_banner(ws: Worksheet, end_col: int) -> None:
    """Extend the Additional Training banner to the current final cert column."""
    banner_start = next(
        (
            col
            for col in range(3, ws.max_column + 1)
            if ws.cell(row=1, column=col).value == "Additional Training"
        ),
        3,
    )
    banner_template = ws.cell(row=1, column=banner_start)

    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_row == 1
            and merged_range.max_row == 1
            and merged_range.min_col == banner_start
        ):
            ws.unmerge_cells(str(merged_range))

    if end_col >= banner_start:
        ws.merge_cells(start_row=1, start_column=banner_start, end_row=1, end_column=end_col)
        banner_cell = ws.cell(row=1, column=banner_start, value="Additional Training")
        banner_cell._style = copy(banner_template._style)


def remove_empty_tracker_header_gaps(ws: Worksheet) -> list[str]:
    """Delete blank Tracker header columns that have no data beneath them."""
    last_col = tracker_last_header_col(ws)
    last_row = max(ws.max_row, 200)
    empty_cols: list[int] = []
    for col in range(3, last_col):
        if ws.cell(row=2, column=col).value not in (None, ""):
            continue
        has_data = any(ws.cell(row=row, column=col).value not in (None, "") for row in range(3, last_row + 1))
        if not has_data:
            empty_cols.append(col)

    if not empty_cols:
        return []

    banner_start = next(
        (
            col
            for col in range(3, ws.max_column + 1)
            if ws.cell(row=1, column=col).value == "Additional Training"
        ),
        3,
    )
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_row == 1
            and merged_range.max_row == 1
            and merged_range.min_col == banner_start
        ):
            ws.unmerge_cells(str(merged_range))

    deleted = []
    for col in sorted(empty_cols, reverse=True):
        deleted.append(get_column_letter(col))
        ws.delete_cols(col, 1)

    refresh_additional_training_banner(ws, tracker_last_header_col(ws))
    return list(reversed(deleted))


def normalize_tracker_column_styles(ws: Worksheet) -> list[str]:
    """Repair cert columns whose headers/data lost the standard Tracker style."""
    template_col = tracker_styled_template_col(ws)
    header_template = ws.cell(row=2, column=template_col)
    data_template = ws.cell(row=3, column=template_col)
    last_col = tracker_last_header_col(ws)
    last_row = max(ws.max_row, 200)
    repaired: list[str] = []

    for col in range(3, last_col + 1):
        header = ws.cell(row=2, column=col)
        if header.value in (None, "") or header.fill.fill_type:
            continue
        header._style = copy(header_template._style)
        ws.column_dimensions[get_column_letter(col)].width = (
            ws.column_dimensions[get_column_letter(template_col)].width or 14
        )
        for row in range(3, last_row + 1):
            cell = ws.cell(row=row, column=col)
            cell._style = copy(data_template._style)
            cell.number_format = "mm/dd/yyyy"
        repaired.append(get_column_letter(col))

    return repaired


def resolve_tracker_cert_column(cert_name: str, cert_columns: dict[str, int]) -> Optional[int]:
    """Find a Tracker column for cert_name using exact, normalized, and
    whitespace-insensitive matching.

    This keeps variants like "OSHA 8Hr Refresher" and "OSHA 8 Hr Refresher"
    pointed at the same workbook column.
    """
    if cert_name in cert_columns:
        return cert_columns[cert_name]
    target = normalize(cert_name)
    target_compact = normalize_compact(cert_name)
    for existing_name, col in cert_columns.items():
        if normalize(existing_name) == target:
            return col
        if normalize_compact(existing_name) == target_compact:
            return col
    return None


def add_tracker_conditional_formatting(
    ws: Worksheet, start_col: int, end_col: int, last_row: int
) -> None:
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    matrix_range = f"{start_letter}3:{end_letter}{last_row}"

    red_fill = PatternFill("solid", fgColor="F8CBAD")
    ws.conditional_formatting.add(
        matrix_range,
        FormulaRule(formula=[f'AND($B3<>"",{start_letter}3="")'], fill=red_fill, stopIfTrue=False),
    )

    orange_fill = PatternFill("solid", fgColor="F4B183")
    expired_formula = (
        f'AND({start_letter}3<>"",'
        f'ISNUMBER({start_letter}3),'
        f'VLOOKUP({start_letter}$2,Certifications!$A:$C,3,FALSE)>0,'
        f'EDATE({start_letter}3,VLOOKUP({start_letter}$2,Certifications!$A:$C,3,FALSE)*12)<TODAY())'
    )
    ws.conditional_formatting.add(
        matrix_range, FormulaRule(formula=[expired_formula], fill=orange_fill, stopIfTrue=False)
    )

    yellow_fill = PatternFill("solid", fgColor="FFE699")
    expiring_formula = (
        f'AND({start_letter}3<>"",'
        f'ISNUMBER({start_letter}3),'
        f'VLOOKUP({start_letter}$2,Certifications!$A:$C,3,FALSE)>0,'
        f'EDATE({start_letter}3,VLOOKUP({start_letter}$2,Certifications!$A:$C,3,FALSE)*12)>=TODAY(),'
        f'EDATE({start_letter}3,VLOOKUP({start_letter}$2,Certifications!$A:$C,3,FALSE)*12)<=TODAY()+60)'
    )
    ws.conditional_formatting.add(
        matrix_range, FormulaRule(formula=[expiring_formula], fill=yellow_fill, stopIfTrue=False)
    )

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    current_formula = f'AND({start_letter}3<>"",ISNUMBER({start_letter}3))'
    ws.conditional_formatting.add(
        matrix_range, FormulaRule(formula=[current_formula], fill=green_fill, stopIfTrue=False)
    )


def apply_renewal_color_rules(ws_tracker: Worksheet) -> None:
    """Replace the Tracker's conditional formatting with 1-year renewal rules.

    A cert tile is colored only when it has a date. Empty cells stay blank.
    Coloring is based on days remaining until the cert's 1-year anniversary
    (cert_date + 12 months):
      - RED    : anniversary already passed (overdue)
      - ORANGE : 30 days or fewer remaining (urgent, still in the future)
      - YELLOW : 31 to 60 days remaining
      - GREEN  : more than 60 days remaining

    Rule order matters because every rule has stopIfTrue=True: RED catches
    expired cells first, so the broader ORANGE rule (<=30) only colors cells
    that are still in the future.

    Applied to a wide static range (C3:AZ200) so future cert columns inherit
    the rules without re-running this helper.
    """
    # Clear existing CF rules in-place so the workbook's dxfs registry is
    # preserved (replacing the whole ConditionalFormattingList orphans the
    # rule fills, which makes Excel render cells without color).
    ws_tracker.conditional_formatting._cf_rules.clear()

    last_row = max(ws_tracker.max_row, 200)
    matrix_range = f"C3:AZ{last_row}"

    # CF differential fills use bgColor (not fgColor) for the actual cell color.
    green_fill = PatternFill(bgColor="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(bgColor="FFE699", fill_type="solid")
    orange_fill = PatternFill(bgColor="F4B084", fill_type="solid")
    red_fill = PatternFill(bgColor="FF0000", fill_type="solid")  # bright red

    # 1) Has a date and anniversary already passed -> RED (expired).
    ws_tracker.conditional_formatting.add(
        matrix_range,
        FormulaRule(
            formula=['AND(C3<>"",ISNUMBER(C3),(EDATE(C3,12)-TODAY())<0)'],
            fill=red_fill,
            stopIfTrue=True,
        ),
    )

    # 2) Has a date and <=30 days remaining -> ORANGE (urgent, still future).
    #    The <0 case is already caught by rule 1 above, so this only matches
    #    the 0..30 future window.
    ws_tracker.conditional_formatting.add(
        matrix_range,
        FormulaRule(
            formula=['AND(C3<>"",ISNUMBER(C3),(EDATE(C3,12)-TODAY())<=30)'],
            fill=orange_fill,
            stopIfTrue=True,
        ),
    )

    # 3) Has a date and 31-60 days remaining -> YELLOW.
    ws_tracker.conditional_formatting.add(
        matrix_range,
        FormulaRule(
            formula=[(
                'AND(C3<>"",ISNUMBER(C3),'
                '(EDATE(C3,12)-TODAY())>30,'
                '(EDATE(C3,12)-TODAY())<=60)'
            )],
            fill=yellow_fill,
            stopIfTrue=True,
        ),
    )

    # 4) Has a date and more than 60 days remaining -> GREEN.
    ws_tracker.conditional_formatting.add(
        matrix_range,
        FormulaRule(
            formula=['AND(C3<>"",ISNUMBER(C3),(EDATE(C3,12)-TODAY())>60)'],
            fill=green_fill,
            stopIfTrue=True,
        ),
    )
    # Empty cells: no rule -> no color (intentional, per user request).


def ensure_tracker_column(ws_tracker: Worksheet, cert_name: str) -> bool:
    """Append cert_name as a new column on the Tracker sheet if missing.

    Inherits styling from the current last column, extends the Additional
    Training banner merge, and adds conditional-formatting rules for just the
    new column. Returns True when a new column was added.
    """
    if cert_name in tracker_cert_columns(ws_tracker):
        return False

    last_existing_col = tracker_last_header_col(ws_tracker)
    template_col = tracker_styled_template_col(ws_tracker)
    header_template = ws_tracker.cell(row=2, column=template_col)
    data_template = ws_tracker.cell(row=3, column=template_col)
    last_row = max(ws_tracker.max_row, 200)

    new_col = last_existing_col + 1
    header_cell = ws_tracker.cell(row=2, column=new_col, value=cert_name)
    header_cell._style = copy(header_template._style)

    width = ws_tracker.column_dimensions[get_column_letter(template_col)].width
    ws_tracker.column_dimensions[get_column_letter(new_col)].width = width or 14

    for row in range(3, last_row + 1):
        cell = ws_tracker.cell(row=row, column=new_col)
        cell._style = copy(data_template._style)
        cell.number_format = "mm/dd/yyyy"

    refresh_additional_training_banner(ws_tracker, new_col)

    add_tracker_conditional_formatting(ws_tracker, new_col, new_col, last_row)
    return True


def ensure_catalog_row(
    ws_certs: Worksheet, cert_name: str, category: str = "Additional Training", validity: int = 0,
) -> bool:
    """Add cert_name to the Certifications sheet if it isn't there yet."""
    existing = {
        normalize(str(row[0]))
        for row in ws_certs.iter_rows(min_row=2, max_col=1, values_only=True)
        if row[0]
    }
    if normalize(cert_name) in existing:
        return False
    row = first_empty_row(ws_certs, key_col=1, start=2, limit=max(ws_certs.max_row + 50, 200))
    ws_certs.cell(row=row, column=1, value=cert_name)
    ws_certs.cell(row=row, column=2, value=category)
    ws_certs.cell(row=row, column=3, value=validity)
    return True


def register_new_certification(
    ws_certs: Worksheet,
    ws_tracker: Worksheet,
    cert_name: str,
    category: str = "Additional Training",
    validity: int = 0,
) -> bool:
    """Add the cert to both Certifications catalog and Tracker columns."""
    added_catalog = ensure_catalog_row(ws_certs, cert_name, category, validity)
    added_column = ensure_tracker_column(ws_tracker, cert_name)
    return added_catalog or added_column


# --- Cross-sheet sync ---
DASHBOARD_DATA_START_ROW = 9  # first row beneath the Compliance sub-header


def get_certifications_list(ws_certs: Worksheet) -> list[tuple[int, str, str]]:
    """Return [(row, name, category)] for every non-empty row in Certifications."""
    certs: list[tuple[int, str, str]] = []
    for r in range(2, ws_certs.max_row + 1):
        name = ws_certs.cell(r, 1).value
        if name in (None, ""):
            continue
        category = ws_certs.cell(r, 2).value or "Additional Training"
        certs.append((r, str(name), str(category)))
    return certs


def _find_tracker_col_for_cert_row(ws_tracker: Worksheet, certs_row: int) -> Optional[int]:
    """Return the Tracker column whose row-2 formula points to Certifications!A{certs_row}."""
    target = certs_row
    for c in range(3, ws_tracker.max_column + 1):
        v = ws_tracker.cell(2, c).value
        if isinstance(v, str):
            m = CERT_FORMULA_RE.match(v)
            if m and int(m.group(1)) == target:
                return c
    return None


def _find_tracker_col_literal(ws_tracker: Worksheet, cert_name: str) -> Optional[int]:
    """Return the Tracker column whose row-2 value is the LITERAL cert_name (no formula)."""
    target_norm = normalize(cert_name)
    for c in range(3, ws_tracker.max_column + 1):
        v = ws_tracker.cell(2, c).value
        if not isinstance(v, str) or v.startswith("="):
            continue
        if normalize(v) == target_norm:
            return c
    return None


def sync_tracker_headers(ws_certs: Worksheet, ws_tracker: Worksheet) -> dict[str, list[str]]:
    """Reconcile Tracker row-2 headers with Certifications.

    Steps:
      1. Any literal Tracker header whose name is not yet in Certifications is
         added to Certifications as Additional Training / validity 0.
      2. For every Certifications row, ensure a Tracker column exists and its
         header is the formula =Certifications!A{row} (propagates renames).
    """
    actions = {"new_certs_from_tracker": [], "new_tracker_cols": [], "converted_to_formula": []}

    # Step 1: literal headers not in the catalog -> register them.
    certs = get_certifications_list(ws_certs)
    known_norm = {normalize(name) for _row, name, _cat in certs}
    for c in range(3, ws_tracker.max_column + 1):
        v = ws_tracker.cell(2, c).value
        if not isinstance(v, str) or not v or v.startswith("="):
            continue
        if normalize(v) in known_norm:
            continue
        new_row = first_empty_row(
            ws_certs, key_col=1, start=2, limit=max(ws_certs.max_row + 50, 200)
        )
        ws_certs.cell(new_row, 1, value=v)
        ws_certs.cell(new_row, 2, value="Additional Training")
        ws_certs.cell(new_row, 3, value=0)
        actions["new_certs_from_tracker"].append(v)

    # Reload cert list after Step 1 additions.
    certs = get_certifications_list(ws_certs)

    # Step 2: for each Certifications row, wire a Tracker column to it.
    for certs_row, name, _cat in certs:
        if _find_tracker_col_for_cert_row(ws_tracker, certs_row) is not None:
            continue
        literal_col = _find_tracker_col_literal(ws_tracker, name)
        if literal_col is not None:
            ws_tracker.cell(2, literal_col, value=f"=Certifications!A{certs_row}")
            actions["converted_to_formula"].append(name)
        else:
            ensure_tracker_column(ws_tracker, name)
            new_col = resolve_tracker_cert_column(name, tracker_cert_columns(ws_tracker))
            if new_col is None:
                new_col = tracker_last_header_col(ws_tracker)
            ws_tracker.cell(2, new_col, value=f"=Certifications!A{certs_row}")
            actions["new_tracker_cols"].append(name)

    return actions


def sync_dashboard_rows(ws_certs: Worksheet, ws_dashboard: Worksheet) -> dict[str, list[str]]:
    """Reconcile Dashboard compliance rows with Certifications.

    Adds a Dashboard row per cert (if missing) and wires name/category cells
    back to Certifications via formulas so renames propagate. The count
    formula uses INDEX/MATCH to find the right Tracker column by name.
    """
    actions = {"added_dashboard_rows": [], "converted_to_formula": []}
    certs = get_certifications_list(ws_certs)

    # Map existing Dashboard rows: by Certifications formula ref, then by literal name.
    by_cert_row: dict[int, int] = {}
    literal_rows: dict[str, int] = {}
    used_rows: set[int] = set()
    end = max(DASHBOARD_DATA_START_ROW, ws_dashboard.max_row) + 5
    for r in range(DASHBOARD_DATA_START_ROW, end):
        v = ws_dashboard.cell(r, 1).value
        if v in (None, ""):
            continue
        used_rows.add(r)
        if isinstance(v, str):
            m = CERT_FORMULA_RE.match(v)
            if m:
                by_cert_row[int(m.group(1))] = r
                continue
            literal_rows[normalize(v)] = r

    def _next_free_row() -> int:
        r = DASHBOARD_DATA_START_ROW
        while r in used_rows:
            r += 1
        used_rows.add(r)
        return r

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for certs_row, name, _cat in certs:
        if certs_row in by_cert_row:
            dashboard_row = by_cert_row[certs_row]
            is_new = False
        elif normalize(name) in literal_rows:
            dashboard_row = literal_rows[normalize(name)]
            is_new = False
            actions["converted_to_formula"].append(name)
        else:
            dashboard_row = _next_free_row()
            is_new = True
            actions["added_dashboard_rows"].append(name)

        r = dashboard_row
        ws_dashboard.cell(r, 1, value=f"=Certifications!A{certs_row}")
        ws_dashboard.cell(r, 2, value=f"=Certifications!B{certs_row}")
        ws_dashboard.cell(r, 3, value=(
            f"=IFERROR(COUNTA(INDEX(Tracker!$C$3:$AZ$200,0,"
            f"MATCH($A{r},Tracker!$C$2:$AZ$2,0))),0)"
        ))
        ws_dashboard.cell(r, 4, value=f"=COUNTA(Workers!A$2:A$200)-C{r}")
        pct = ws_dashboard.cell(r, 5, value=f"=IFERROR(C{r}/(C{r}+D{r}),0)")
        pct.number_format = "0.0%"
        bar = ws_dashboard.cell(r, 6, value=f'=REPT("■",ROUND(E{r}*20,0))')
        if is_new:
            bar.font = Font(color="2E7D32", size=11)
            for col in range(1, 7):
                ws_dashboard.cell(r, col).border = border

    return actions


def _replace_tracker_table_with_autofilter(ws_tracker: Worksheet) -> bool:
    """Drop the formal Excel Table on the Tracker and use AutoFilter instead.

    An Excel Table requires its ref and tableColumns metadata to stay in sync
    with the actual data range. When we add cert columns dynamically that
    metadata gets stale and Excel flags the workbook as corrupted ("We found a
    problem with some content..."). AutoFilter gives users the same
    column-header filter dropdowns without the fragility.
    """
    changed = False
    if "TrackerTable" in ws_tracker.tables:
        del ws_tracker.tables["TrackerTable"]
        changed = True

    last_row = max(ws_tracker.max_row, 200)
    last_col = max(tracker_last_header_col(ws_tracker), 26)
    new_ref = f"A2:{get_column_letter(last_col)}{last_row}"
    if ws_tracker.auto_filter.ref != new_ref:
        ws_tracker.auto_filter.ref = new_ref
        changed = True
    return changed


def _resync_certifications_table(ws_certs: Worksheet) -> bool:
    """Keep the CertificationsTable ref covering all populated cert rows.

    Adding new certs grows the sheet downward; if the Table's ref is shorter
    than the data, Excel can flag the file. We just stretch the existing ref to
    include the current populated range (no column changes needed since
    Certifications has fixed columns A-D).
    """
    name = "CertificationsTable"
    if name not in ws_certs.tables:
        return False
    tbl = ws_certs.tables[name]
    try:
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
    except Exception:
        return False

    target_last_row = max(ws_certs.max_row, max_row, 60)
    if target_last_row <= max_row:
        return False
    tbl.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{target_last_row}"
    )
    return True


def sync_workbook(wb) -> dict[str, list[str]]:
    """Reconcile Certifications, Tracker, and Dashboard so they share names.

    Certifications is the source of truth. Tracker row-2 headers and Dashboard
    columns A/B become formulas referencing Certifications, so renames
    propagate automatically.
    """
    ws_certs = wb["Certifications"]
    ws_tracker = wb["Tracker"]
    ws_dashboard = wb["Dashboard"]

    result: dict[str, list[str]] = {}
    removed_gaps = remove_empty_tracker_header_gaps(ws_tracker)
    if removed_gaps:
        result["tracker.removed_blank_header_columns"] = removed_gaps

    repaired_columns = normalize_tracker_column_styles(ws_tracker)
    if repaired_columns:
        result["tracker.repaired_column_styles"] = repaired_columns

    tracker_actions = sync_tracker_headers(ws_certs, ws_tracker)
    dashboard_actions = sync_dashboard_rows(ws_certs, ws_dashboard)

    # After any structural changes, keep Excel Tables consistent so the file
    # opens cleanly without "we found a problem with some content..." prompts.
    if _replace_tracker_table_with_autofilter(ws_tracker):
        result["tracker.table_swapped_for_autofilter"] = ["TrackerTable removed; AutoFilter set"]
    if _resync_certifications_table(ws_certs):
        result["certifications.table_extended"] = [
            f"CertificationsTable ref now covers row {ws_certs.max_row}"
        ]

    # Re-apply the 1-year renewal color rules over the whole Tracker matrix.
    apply_renewal_color_rules(ws_tracker)
    result["tracker.renewal_rules_applied"] = [
        ">60d=green | 31-60d=yellow | <=30d=orange | past=red | empty=blank"
    ]

    # Prefix keys so Tracker and Dashboard actions stay separate even when they
    # share a name like "converted_to_formula".
    for key, items in tracker_actions.items():
        if items:
            result[f"tracker.{key}"] = list(items)
    for key, items in dashboard_actions.items():
        if items:
            result[f"dashboard.{key}"] = list(items)
    return result


def clean_cert_header(raw: str) -> str:
    """Normalize a raw PDF cert header into a usable catalog name."""
    cleaned = raw.replace("\n", " ")
    cleaned = re.sub(r"([A-Za-z])(\d)", r"\1 \2", cleaned)
    cleaned = re.sub(r"(\d)([A-Za-z])", r"\1 \2", cleaned)
    # OSHA course codes often come in as "OSHA 30511" (spaces lost between the
    # level and the course number). Split those into "OSHA 30 511".
    cleaned = re.sub(r"\bOSHA\s+(\d{2})(\d{3})\b", r"OSHA \1 \2", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def ensure_required_certifications(ws_certs: Worksheet, ws_tracker: Worksheet) -> list[str]:
    existing_cert_names = {
        normalize_compact(str(row[0])): row[0]
        for row in ws_certs.iter_rows(min_row=2, max_col=1, values_only=True)
        if row[0]
    }
    added = []
    for cert_name, category, validity in REQUIRED_BASELINE_CERTS:
        if normalize_compact(cert_name) in existing_cert_names:
            continue
        row = first_empty_row(ws_certs, key_col=1, start=2, limit=max(ws_certs.max_row + 50, 200))
        ws_certs.cell(row=row, column=1, value=cert_name)
        ws_certs.cell(row=row, column=2, value=category)
        ws_certs.cell(row=row, column=3, value=validity)
        added.append(cert_name)
        existing_cert_names[normalize_compact(cert_name)] = cert_name

    current_columns = tracker_cert_columns(ws_tracker)
    current_column_keys = {normalize_compact(name) for name in current_columns}
    missing_headers = [
        name
        for name, _category, _validity in REQUIRED_BASELINE_CERTS
        if normalize_compact(name) not in current_column_keys
    ]
    if not missing_headers:
        return added

    last_existing_col = tracker_last_header_col(ws_tracker)
    template_col = tracker_styled_template_col(ws_tracker)
    header_template = ws_tracker.cell(row=2, column=template_col)
    data_template = ws_tracker.cell(row=3, column=template_col)
    last_row = max(ws_tracker.max_row, 200)

    for offset, cert_name in enumerate(missing_headers, start=1):
        col = last_existing_col + offset
        header_cell = ws_tracker.cell(row=2, column=col, value=cert_name)
        header_cell._style = copy(header_template._style)

        width = ws_tracker.column_dimensions[get_column_letter(template_col)].width
        ws_tracker.column_dimensions[get_column_letter(col)].width = width or 14

        for row in range(3, last_row + 1):
            cell = ws_tracker.cell(row=row, column=col)
            cell._style = copy(data_template._style)
            cell.number_format = "mm/dd/yyyy"

    last_added_col = last_existing_col + len(missing_headers)
    refresh_additional_training_banner(ws_tracker, last_added_col)

    add_tracker_conditional_formatting(
        ws_tracker,
        last_existing_col + 1,
        last_added_col,
        last_row,
    )
    return list(dict.fromkeys(added + missing_headers))


def find_tracker_row(
    ws: Worksheet, contractor: str, worker: str
) -> Optional[int]:
    c_norm = normalize_company(contractor)
    w_norm = normalize(worker)
    for r in range(3, 2000):
        wval = ws.cell(row=r, column=2).value
        if wval in (None, ""):
            return None
        cval = ws.cell(row=r, column=1).value or ""
        if normalize(str(wval)) == w_norm and normalize_company(str(cval)) == c_norm:
            return r
    return None


# --- Main import logic ---
def import_pdf(pdf_path: Path) -> dict:
    contractor, primary_contact, workers_data = extract_pdf_data(pdf_path)
    if not contractor:
        # Dump first 400 chars of extracted text so we can fix the regex
        with pdfplumber.open(pdf_path) as pdf:
            snippet = (pdf.pages[0].extract_text() or "")[:400]
        raise RuntimeError(
            f"Could not find contractor name in {pdf_path.name}.\n"
            f"--- First 400 chars of extracted text ---\n{snippet}\n"
            f"--- end ---"
        )
    if not workers_data:
        raise RuntimeError(f"No worker rows extracted from {pdf_path.name}")

    wb = load_workbook(WORKBOOK_PATH)
    ws_contractors = wb["Contractors"]
    ws_workers = wb["Workers"]
    ws_certs = wb["Certifications"]
    ws_tracker = wb["Tracker"]

    added_cert_columns = ensure_required_certifications(ws_certs, ws_tracker)

    alias_map = build_cert_alias_map(ws_certs)
    cert_columns = tracker_cert_columns(ws_tracker)

    # Auto-register any cert header in the PDF that the catalog doesn't know yet.
    # Defaults: Additional Training, validity=0 (no expiration). The user can
    # edit the validity on the Certifications sheet later.
    pdf_headers: set[str] = set()
    for worker_certs in workers_data.values():
        pdf_headers.update(worker_certs.keys())

    # Pre-build a whitespace-insensitive view of the alias map so headers
    # like "OSHA 8 Hr Refresher" still resolve to canonical "OSHA 8Hr Refresher".
    compact_alias_map = {normalize_compact(k): v for k, v in alias_map.items()}
    # Also fold every existing cert column into the compact map directly, so
    # a column header that's already in the Tracker but missing from the
    # Certifications-derived alias map still counts as "known".
    for canonical_name in cert_columns:
        compact_alias_map.setdefault(normalize_compact(canonical_name), canonical_name)

    def _exact_known(name: str) -> bool:
        """Is `name` an exact (normalized) match for a cert already in the catalog?
        Deliberately does NOT use the fuzzy token-overlap fallback, so a new cert
        like 'OSHA 30 501' does not get swallowed by existing 'OSHA 30'.
        Whitespace-insensitive so 'OSHA 8 Hr Refresher' matches 'OSHA 8Hr Refresher'.
        """
        key = normalize(name)
        if not key:
            return False
        canonical = alias_map.get(key)
        if canonical is None:
            canonical = alias_map.get(normalize(re.sub(r"[()]", "", key)))
        if canonical is None:
            canonical = compact_alias_map.get(normalize_compact(name))
        return canonical is not None and canonical in cert_columns

    auto_added: list[str] = []
    for header in sorted(pdf_headers):
        if _exact_known(header):
            continue
        cleaned = clean_cert_header(header)
        if not cleaned or _exact_known(cleaned):
            continue
        if register_new_certification(ws_certs, ws_tracker, cleaned):
            auto_added.append(cleaned)

    # Sync Certifications <-> Tracker <-> Dashboard so any cert added in any
    # sheet appears in all three, and Tracker headers + Dashboard name/category
    # cells become formulas that track Certifications renames automatically.
    sync_actions = sync_workbook(wb)

    if auto_added or sync_actions:
        alias_map = build_cert_alias_map(ws_certs)
        cert_columns = tracker_cert_columns(ws_tracker)

    stats = {
        "contractor": contractor,
        "primary_contact": primary_contact,
        "contractor_added": False,
        "contact_added": False,
        "workers_added": 0,
        "workers_existing": 0,
        "certs_updated": 0,
        "certs_unchanged": 0,
        "cert_columns_added": added_cert_columns,
        "new_certs_added": auto_added,
        "sync_actions": sync_actions,
        "unmatched_headers": set(),
    }

    # 1) Contractor + primary contact
    # Use a punctuation-insensitive match so "GeoEnviroTech, Inc." (workbook)
    # and "GeoEnviroTech, Inc" (PDF, period stripped by extract_contractor)
    # are recognized as the same company.
    c_norm = normalize_company(contractor)
    c_row = None
    for r in range(2, 2000):
        val = ws_contractors.cell(row=r, column=1).value
        if val in (None, ""):
            break
        if normalize_company(str(val)) == c_norm:
            c_row = r
            # Reuse the workbook's spelling so downstream sheets stay
            # consistent with what's already there.
            contractor = str(val)
            break
    if c_row is None:
        c_row = first_empty_row(ws_contractors, key_col=1)
        ws_contractors.cell(row=c_row, column=1, value=contractor)
        stats["contractor_added"] = True
    stats["contractor"] = contractor

    # Fill primary contact if extracted and currently blank
    if primary_contact:
        contact_cell = ws_contractors.cell(row=c_row, column=2)
        if not contact_cell.value:
            contact_cell.value = primary_contact
            stats["contact_added"] = True

    # 2) Workers and tracker rows
    for worker, certs in workers_data.items():
        w_norm = normalize(worker)

        # Add worker to Workers sheet if not already present (under ANY contractor)
        existing_row = None
        for r in range(2, 2000):
            nm = ws_workers.cell(row=r, column=1).value
            if nm in (None, ""):
                break
            if normalize(str(nm)) == w_norm:
                existing_row = r
                break

        if existing_row is None:
            new_row = first_empty_row(ws_workers, key_col=1)
            ws_workers.cell(row=new_row, column=1, value=worker)
            ws_workers.cell(row=new_row, column=2, value=contractor)
            ws_workers.cell(row=new_row, column=4, value="active")
            stats["workers_added"] += 1
        else:
            stats["workers_existing"] += 1

        # Find or create tracker row for (contractor, worker)
        t_row = find_tracker_row(ws_tracker, contractor, worker)
        if t_row is None:
            t_row = first_empty_row(ws_tracker, key_col=2, start=3)
            ws_tracker.cell(row=t_row, column=1, value=contractor)
            ws_tracker.cell(row=t_row, column=2, value=worker)

        # Update cert dates (newer wins)
        for header_name, dt in certs.items():
            canonical = match_cert(header_name, alias_map)
            col = resolve_tracker_cert_column(canonical, cert_columns) if canonical else None
            if not canonical or col is None:
                stats["unmatched_headers"].add(header_name)
                continue
            cell = ws_tracker.cell(row=t_row, column=col)
            existing = cell.value
            if isinstance(existing, datetime):
                existing = existing.date()
            if existing is None or (isinstance(existing, date) and dt > existing):
                cell.value = dt
                cell.number_format = "mm/dd/yyyy"
                stats["certs_updated"] += 1
            else:
                stats["certs_unchanged"] += 1

    wb.save(WORKBOOK_PATH)
    return stats


def _run_sync_only() -> int:
    if not WORKBOOK_PATH.exists():
        print(f"Workbook not found at {WORKBOOK_PATH}.")
        return 2
    try:
        wb = load_workbook(WORKBOOK_PATH)
    except PermissionError:
        print(
            "[ERROR] Permission denied — the Excel file is open.\n"
            f"        Close '{WORKBOOK_PATH.name}' in Excel and try again."
        )
        return 2

    actions = sync_workbook(wb)
    wb.save(WORKBOOK_PATH)
    print(f"Sync complete: {WORKBOOK_PATH}")
    if not actions:
        print("  (no changes needed — everything already in sync)")
        return 0
    for key, items in actions.items():
        if not items:
            continue
        label = key.replace("_", " ").capitalize()
        print(f"  {label}:")
        for item in items:
            print(f"    - {item}")
    return 0


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("Usage: python import_pdf.py <file.pdf> [more.pdf ...]")
        print("       python import_pdf.py --sync   # reconcile sheets only")
        return 1
    if "--sync" in argv[1:]:
        return _run_sync_only()
    if not WORKBOOK_PATH.exists():
        print(f"Workbook not found at {WORKBOOK_PATH}. Run tools/build_cert_tracker.py first.")
        return 2

    any_success = False
    for arg in argv[1:]:
        pdf_path = Path(arg)
        if not pdf_path.exists():
            print(f"[SKIP] Not found: {pdf_path}")
            continue
        print(f"\n=== Importing: {pdf_path.name} ===")
        try:
            stats = import_pdf(pdf_path)
        except PermissionError:
            print(
                "[ERROR] Permission denied — the Excel file is open.\n"
                "        Close 'Contractor Certifications Tracker.xlsx' in Excel and try again."
            )
            continue
        except Exception as exc:
            print(f"[ERROR] {exc}")
            continue
        any_success = True
        print(f"  Contractor:        {stats['contractor']}"
              f"{' (NEW)' if stats['contractor_added'] else ''}")
        if stats["primary_contact"]:
            print(f"  Primary contact:   {stats['primary_contact']}"
                  f"{' (saved)' if stats['contact_added'] else ''}")
        else:
            print("  Primary contact:   (not detected in PDF)")
        print(f"  Workers added:     {stats['workers_added']}")
        print(f"  Workers existing:  {stats['workers_existing']}")
        if stats["cert_columns_added"]:
            print(f"  Cert columns added:{' ' if len(stats['cert_columns_added']) < 10 else ''}{', '.join(stats['cert_columns_added'])}")
        if stats.get("new_certs_added"):
            print(f"  New certs auto-added (Additional Training, 0-yr validity):")
            for name in stats["new_certs_added"]:
                print(f"       + {name}")
        print(f"  Cert dates set:    {stats['certs_updated']}")
        print(f"  Cert dates kept:   {stats['certs_unchanged']} (already newer or same)")
        if stats["unmatched_headers"]:
            print("  [!] Unmatched cert columns (not in Certifications sheet):")
            for h in sorted(stats["unmatched_headers"]):
                print(f"       - {h}")

    if any_success:
        print(f"\nSaved: {WORKBOOK_PATH}")
    else:
        print("\nNo changes saved.")
    return 0 if any_success else 3


if __name__ == "__main__":
    sys.exit(main(sys.argv))
