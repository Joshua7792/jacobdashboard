"""Consolidate duplicate workers and duplicate cert columns left over from
earlier imports that used stricter matching.

What it does:
- Workers/Tracker: rows with the same (contractor mod punctuation, worker)
  are merged into the FIRST occurrence; cert dates take the newer value.
  Duplicate rows are deleted.
- Cert columns: columns whose headers differ only in whitespace
  (e.g. "OSHA 8 Hr Refresher" vs "OSHA 8Hr Refresher") are merged into
  the FIRST occurrence; the duplicate column is deleted.
- Certifications sheet: duplicate cert rows (whitespace-insensitive) are
  deleted; the survivor's name is canonical.
- Contractors: rows that match modulo trailing punctuation are merged.

Run with the workbook closed in Excel:
    python cert_tracker/scripts/dedupe_workbook.py
"""
from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Reuse normalization helpers from import_pdf so we stay consistent.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from import_pdf import (  # noqa: E402
    CERT_FORMULA_RE,
    WORKBOOK_PATH,
    normalize,
    normalize_company,
    normalize_compact,
    sync_workbook,
)


def _date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def dedupe_certifications(ws_certs) -> list[str]:
    """Drop Certifications rows whose name (whitespace-insensitive) duplicates
    an earlier row. Returns the deleted names so the caller can remap Tracker
    columns."""
    seen: dict[str, int] = {}
    rows_to_delete: list[tuple[int, str]] = []
    for r in range(2, ws_certs.max_row + 1):
        name = ws_certs.cell(r, 1).value
        if not name:
            continue
        key = normalize_compact(str(name))
        if not key:
            continue
        if key in seen:
            rows_to_delete.append((r, str(name)))
        else:
            seen[key] = r
    # Delete from bottom up so row numbers don't shift mid-loop.
    deleted = []
    for r, name in sorted(rows_to_delete, reverse=True):
        ws_certs.delete_rows(r, 1)
        deleted.append(name)
    return deleted


def _resolve_tracker_header(ws_tracker, ws_certs, col: int) -> str:
    """Return the canonical cert name for a Tracker column, resolving formula
    headers like =Certifications!A17 to the actual cert name."""
    val = ws_tracker.cell(2, col).value
    if isinstance(val, str):
        m = CERT_FORMULA_RE.match(val)
        if m:
            certs_row = int(m.group(1))
            resolved = ws_certs.cell(certs_row, 1).value
            if resolved:
                return str(resolved)
    return str(val) if val is not None else ""


def dedupe_tracker_columns(ws_tracker, ws_certs) -> list[str]:
    """Merge Tracker cert columns whose headers (whitespace-insensitive) collide.
    Surviving column gets the newer date for any conflicting cell. Duplicate
    columns are deleted."""
    # First pass: map compact-key -> first surviving column index.
    survivors: dict[str, int] = {}
    duplicates: list[tuple[int, str, int]] = []  # (col, compact_key, survivor_col)
    for col in range(3, ws_tracker.max_column + 1):
        name = _resolve_tracker_header(ws_tracker, ws_certs, col)
        key = normalize_compact(name)
        if not key:
            continue
        if key in survivors:
            duplicates.append((col, key, survivors[key]))
        else:
            survivors[key] = col

    if not duplicates:
        return []

    last_row = max(ws_tracker.max_row, 200)
    deleted_headers = []

    # Merge dup column data into the survivor (newer date wins).
    for dup_col, _key, survivor_col in duplicates:
        for r in range(3, last_row + 1):
            dup_val = _date(ws_tracker.cell(r, dup_col).value)
            if dup_val is None:
                continue
            surv_cell = ws_tracker.cell(r, survivor_col)
            surv_val = _date(surv_cell.value)
            if surv_val is None or dup_val > surv_val:
                surv_cell.value = dup_val
                surv_cell.number_format = "mm/dd/yyyy"
        deleted_headers.append(_resolve_tracker_header(ws_tracker, ws_certs, dup_col))

    # Delete duplicate columns from rightmost to leftmost so indices don't shift.
    for col, _, _ in sorted(duplicates, key=lambda t: t[0], reverse=True):
        ws_tracker.delete_cols(col, 1)

    return deleted_headers


def dedupe_tracker_rows(ws_tracker) -> int:
    """Merge Tracker rows that share (contractor mod punctuation, worker).
    Newer cert date wins. Returns the number of rows deleted."""
    survivors: dict[tuple[str, str], int] = {}
    duplicates: list[tuple[int, int]] = []  # (dup_row, survivor_row)
    for r in range(3, ws_tracker.max_row + 1):
        worker = ws_tracker.cell(r, 2).value
        if worker in (None, ""):
            continue
        contractor = ws_tracker.cell(r, 1).value or ""
        key = (normalize_company(str(contractor)), normalize(str(worker)))
        if not key[1]:
            continue
        if key in survivors:
            duplicates.append((r, survivors[key]))
        else:
            survivors[key] = r

    last_col = ws_tracker.max_column
    for dup_row, survivor_row in duplicates:
        for col in range(3, last_col + 1):
            dup_val = _date(ws_tracker.cell(dup_row, col).value)
            if dup_val is None:
                continue
            surv_cell = ws_tracker.cell(survivor_row, col)
            surv_val = _date(surv_cell.value)
            if surv_val is None or dup_val > surv_val:
                surv_cell.value = dup_val
                surv_cell.number_format = "mm/dd/yyyy"

    for dup_row, _ in sorted(duplicates, key=lambda t: t[0], reverse=True):
        ws_tracker.delete_rows(dup_row, 1)

    return len(duplicates)


def dedupe_workers_sheet(ws_workers) -> int:
    """Drop Workers rows that share a normalized name. Keep the first."""
    seen: set[str] = set()
    rows_to_delete: list[int] = []
    for r in range(2, ws_workers.max_row + 1):
        nm = ws_workers.cell(r, 1).value
        if nm in (None, ""):
            continue
        key = normalize(str(nm))
        if key in seen:
            rows_to_delete.append(r)
        else:
            seen.add(key)
    for r in sorted(rows_to_delete, reverse=True):
        ws_workers.delete_rows(r, 1)
    return len(rows_to_delete)


def dedupe_contractors(ws_contractors) -> int:
    """Drop Contractors rows that match modulo trailing punctuation."""
    seen: dict[str, int] = {}
    rows_to_delete: list[int] = []
    for r in range(2, ws_contractors.max_row + 1):
        nm = ws_contractors.cell(r, 1).value
        if nm in (None, ""):
            continue
        key = normalize_company(str(nm))
        if key in seen:
            # Backfill missing primary contact onto the survivor before deleting.
            survivor = seen[key]
            surv_cell = ws_contractors.cell(survivor, 2)
            dup_cell = ws_contractors.cell(r, 2)
            if not surv_cell.value and dup_cell.value:
                surv_cell.value = dup_cell.value
            rows_to_delete.append(r)
        else:
            seen[key] = r
    for r in sorted(rows_to_delete, reverse=True):
        ws_contractors.delete_rows(r, 1)
    return len(rows_to_delete)


def main() -> int:
    if not WORKBOOK_PATH.exists():
        print(f"Workbook not found: {WORKBOOK_PATH}")
        return 2
    try:
        wb = load_workbook(WORKBOOK_PATH)
    except PermissionError:
        print(f"[ERROR] Workbook is open in Excel. Close it and re-run.")
        return 2

    ws_certs = wb["Certifications"]
    ws_tracker = wb["Tracker"]
    ws_workers = wb["Workers"]
    ws_contractors = wb["Contractors"]

    print(f"Deduplicating: {WORKBOOK_PATH}")

    # Order matters: merge Tracker columns FIRST (while formula headers still
    # resolve correctly via Certifications rows), then dedupe Certifications.
    deleted_cols = dedupe_tracker_columns(ws_tracker, ws_certs)
    if deleted_cols:
        print(f"  Tracker duplicate columns merged: {deleted_cols}")

    deleted_certs = dedupe_certifications(ws_certs)
    if deleted_certs:
        print(f"  Certifications duplicates removed: {deleted_certs}")

    deleted_rows = dedupe_tracker_rows(ws_tracker)
    if deleted_rows:
        print(f"  Tracker duplicate worker rows merged: {deleted_rows}")

    deleted_workers = dedupe_workers_sheet(ws_workers)
    if deleted_workers:
        print(f"  Workers duplicate rows removed: {deleted_workers}")

    deleted_contractors = dedupe_contractors(ws_contractors)
    if deleted_contractors:
        print(f"  Contractors duplicate rows removed: {deleted_contractors}")

    # Re-run the cross-sheet sync so Tracker headers and Dashboard rows
    # rewire to the (possibly renumbered) Certifications rows.
    sync_actions = sync_workbook(wb)
    if sync_actions:
        print("  Sync after dedupe:")
        for k, v in sync_actions.items():
            print(f"    {k}: {v}")

    wb.save(WORKBOOK_PATH)
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
